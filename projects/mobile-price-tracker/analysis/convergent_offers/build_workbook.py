# -*- coding: utf-8 -*-
"""CODE TASK #26 — Convergent-offers snapshot (TIM Ultracombo / Vivo Total / Claro Multi), SP, 2026-07.

Builds `convergent_offers_2026-07.xlsx` — one English-format sheet per carrier plus Notes & Sources.
Data provenance (all captures in ./raw/, kept local):
  - TIM: TIM's published Ultracombo table (transcribed by the Bridge), verified against the
    Teletime article of 2026-07-16 (raw/teletime_ultracombo.html). TIM's own site had no
    dedicated Ultracombo page on 2026-07-17 (three URL probes -> 404).
  - Vivo: raw/vivo_total_rendered.html — Playwright render of the Vivo Total combos page,
    SP confirmed via the "São Paulo (SP)" location label. 11 combo cards (.unique-card).
  - Claro: raw/claro_multi_rendered.html — Playwright render of claro.com.br/multi,
    "Fibra + Móvel" tab; SP segmentation confirmed in __NEXT_DATA__ (uf=SP, not user-changed).
    5 combos rendered (a 6th exists in the embedded JSON but is not shown to SP visitors).

Prices are numeric cells; the TIM "*" marker is applied via the number format so the cell
stays numeric. Run: python build_workbook.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "convergent_offers_2026-07.xlsx"

SNAPSHOT_DATE = "2026-07-17"

# House style (CONTEXT §7)
HEADER_FILL = "1F3864"
BAND_FILL = "F2F5FA"
TAB_COLORS = {"TIM": "0033A0", "Vivo": "660099", "Claro": "DA291C", "Notes & Sources": "8A8A8A"}

COLUMNS = ["Mobile Plan", "Home Internet", "Streaming", "Other Benefits", "Monthly Price (R$)"]
WIDTHS = [46, 30, 44, 78, 18]

PRICE_FMT = '"R$" #,##0.00'
PRICE_FMT_STAR = '"R$" #,##0.00"*"'

# ---------------------------------------------------------------------------
# TIM — given table (TIM's published Ultracombo grid), verified vs Teletime.
# star=True -> the row carries TIM's "*" marker (SP-only offer; see Notes).
TIM_BENEFIT = "Higher TIM Mais loyalty-program tier included (level varies by package — per Teletime)"
TIM_ROWS = [
    ("TIM Black Light (65GB)*", "TIM Ultrafibra 500 Mbps*", "N/A", TIM_BENEFIT, 89.99, True),
    ("TIM Black (70GB)", "TIM Ultrafibra 600 Mbps", "Globoplay", TIM_BENEFIT, 149.99, False),
    ("TIM Black Premium (115GB)", "TIM Ultrafibra 600 Mbps", "Globoplay", TIM_BENEFIT, 169.99, False),
    ("TIM Black Família VIP (200GB)", "TIM Ultrafibra 600 Mbps", "Globoplay", TIM_BENEFIT, 314.99, False),
    ("TIM Black (70GB)*", "TIM Ultrafibra 1 Gbps*", "Paramount+", TIM_BENEFIT, 139.99, True),
    ("TIM Black Premium (115GB)*", "TIM Ultrafibra 1 Gbps*", "Paramount+", TIM_BENEFIT, 169.99, True),
    ("TIM Black Família VIP (200GB)*", "TIM Ultrafibra 1 Gbps*", "Paramount+", TIM_BENEFIT, 314.99, True),
]

# ---------------------------------------------------------------------------
# Vivo — parsed from raw/vivo_total_rendered.html (11 .unique-card combos).
def _vivo_perks(travel, gemini_months, prime="Amazon Prime + Apple Music free for 6 months (then R$ 23.90/mo)"):
    return (
        f"{travel} roaming; Gemini AI Plus free for {gemini_months} months (then R$ 24.99/mo); "
        f"{prime}; Waze & Moovit browsing is data-free"
    )

VIVO_ROWS = [
    ("Vivo Total Ultra — Vivo Pós 70 GB", "Vivo Fibra 700 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas", 6), 170, False),
    ("Vivo Total Pro — Vivo Pós 60 GB (50 GB + 10 GB bonus for 1 year)", "Vivo Fibra 500 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas", 6,
                 "Amazon Prime + Apple Music free for 6 months (then R$ 28.90/mo)"), 160, False),
    ("Vivo Total Família 2 — Vivo Pós 120 GB + 1 additional line", "Vivo Fibra 700 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 270, False),
    ("Vivo Total Família 3 — Vivo Pós 180 GB + 2 additional lines", "Vivo Fibra 700 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 330, False),
    ("Vivo Total Família 4 — Vivo Pós 240 GB + 3 additional lines", "Vivo Fibra 700 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 420, False),
    ("Vivo Total Família 5 — Vivo Pós 300 GB + 4 additional lines", "Vivo Fibra 700 Mbps", "N/A",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 520, False),
    ("Vivo Total V 1 Giga — Vivo Pós 600 GB + 10 additional lines", "Vivo Fibra 1 Gbps",
     "Vivo TV Completo (140+ channels, HBO Max Standard + Telecine, 4 TV devices)",
     _vivo_perks("Vivo Travel Mundo", 12, "Amazon Prime free for 6 months (then R$ 13.90/mo)"), 1200, False),
    ("Vivo Total Ultra + TV online — Vivo Pós 70 GB", "Vivo Fibra 700 Mbps",
     "Vivo TV Estendido online (80+ channels, up to 3 simultaneous screens)",
     _vivo_perks("Vivo Travel Américas", 6), 190, False),
    ("Vivo Total Ultra + TV — Vivo Pós 70 GB", "Vivo Fibra 700 Mbps",
     "Vivo TV Avançado (120+ channels, 2 TV devices + Vivo TV Online)",
     _vivo_perks("Vivo Travel Américas", 6), 290, False),
    ("Vivo Família 2 (+ TV online) — Vivo Pós 120 GB + 1 additional line", "Vivo Fibra 700 Mbps",
     "Vivo TV Estendido online (80+ channels, up to 3 simultaneous screens)",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 290, False),
    ("Vivo Família 2 (+ TV) — Vivo Pós 120 GB + 1 additional line", "Vivo Fibra 700 Mbps",
     "Vivo TV Avançado (120+ channels, 2 TV devices + Vivo TV Online)",
     _vivo_perks("Vivo Travel Américas & Europe", 12), 390, False),
]

# ---------------------------------------------------------------------------
# Claro — parsed from raw/claro_multi_rendered.html ("Fibra + Móvel" tab, 5 rendered combos).
CLARO_CLOUD = "cloud storage included (Google One or iCloud+)"
CLARO_WIFI = "Wi-Fi router + free installation"
CLARO_ROWS = [
    ("Claro Controle 41GB (35 GB plan + 6 GB bonus)", "Claro fibra 350 Mbps", "Globoplay",
     f"WhatsApp unlimited; extra data bonus on apps (YouTube, Instagram, Facebook, TikTok, X, Claro tv+); {CLARO_WIFI}",
     129.80, False),
    ("Claro Pós 60GB", "Claro fibra 600 Mbps", "Globoplay",
     f"Passaporte Américas roaming included; WhatsApp unlimited; {CLARO_CLOUD}; {CLARO_WIFI}",
     159.90, False),
    ("Claro Pós 60GB", "Claro fibra 1 Gbps (Wi-Fi 6)", "Globoplay",
     f"Passaporte Américas roaming included; WhatsApp unlimited; {CLARO_CLOUD}; Wi-Fi 6 router + free installation",
     209.90, False),
    ("Claro Pós 150GB", "Claro fibra 1 Gbps (Wi-Fi 6)", "Globoplay",
     f"Passaporte Américas + Europe roaming; 1 additional line included; {CLARO_CLOUD}; Wi-Fi 6 router + free installation",
     329.90, False),
    ("Claro Pós 200GB", "Claro fibra 1 Gbps (Wi-Fi 6)", "Globoplay",
     f"Passaporte Mundo roaming; 2 dependent lines included; {CLARO_CLOUD}; Wi-Fi 6 router + free installation",
     389.90, False),
]

# ---------------------------------------------------------------------------
NOTES = [
    ("SNAPSHOT", ""),
    ("Snapshot date", f"{SNAPSHOT_DATE} (one-pass capture). Region: São Paulo — SP capital default segmentation on all three sites."),
    ("Scope", "Convergent (mobile + fiber) bundles as publicly listed by each carrier. Every row is sourced from the captures below; nothing inferred."),
    ("", ""),
    ("SOURCES", ""),
    ("TIM", "TIM's published Ultracombo table (offer launched 2026-07-16), verified against Teletime: "
            "https://teletime.com.br/16/07/2026/tim-ultracombo-40/ (raw/teletime_ultracombo.html). "
            "TIM's own site had NO dedicated Ultracombo page on 2026-07-17 — probes of "
            "tim.com.br/sp/para-voce/planos/ultracombo, /sp/ultracombo and /ultracombo all returned 404."),
    ("Vivo", "https://vivo.com.br/para-voce/produtos-e-servicos/combos/vivo-total — headless-browser render "
             "(raw/vivo_total_rendered.html). SP confirmed via the page's \"São Paulo (SP)\" location label. 11 combo cards."),
    ("Claro", "https://www.claro.com.br/multi — headless-browser render, \"Fibra + Móvel\" tab "
              "(raw/claro_multi_rendered.html + raw/claro_multi_next_data.json). SP segmentation confirmed in the page JSON "
              "(uf=SP, São Paulo, DDD 11, not user-changed). 5 combos rendered."),
    ("", ""),
    ("FOOTNOTES (TIM)", ""),
    ("* marker", "\"Ofertas disponíveis apenas em São Paulo\" = offer available ONLY in São Paulo (state). "
                 "Confirmed from the caption of TIM's own table as reproduced in the Teletime article — not guessed."),
    ("Payment condition", "All TIM prices shown are valid for automatic-debit payment (\"débito automático\"), per the same caption. "
                          "Offers vary by region and TIM Ultrafibra availability."),
    ("TIM claims", "Up to 40% savings vs contracting the same services separately (TIM's example: TIM Black Família VIP 200GB + "
                   "Ultrafibra 600 Mbps + Globoplay = R$ 314.99/mo vs ~R$ 570 separately). Packages also place the customer in "
                   "higher TIM Mais loyalty-program tiers; the tier per package is not published."),
    ("Streaming brand", "The transcribed TIM table wrote \"Paramount\"; the Teletime article states \"Paramount+\" — the sheet uses Paramount+."),
    ("", ""),
    ("CAVEATS (Vivo)", ""),
    ("Promo window", "\"Vivo Total Ultra\" is tagged \"POR TEMPO LIMITADO\" (limited time) and \"Exclusivo no site\" (website-exclusive); "
                     "the Ultra + TV variants are also website-exclusive."),
    ("Installation", "Every card states: free installation conditional on a loyalty commitment (\"mediante fidelização\") and "
                     "Wi-Fi bonus conditional on payments in good standing (\"mediante adimplência\")."),
    ("Streaming add-on (NOT included)", "The 6-app streaming pack (Netflix, Globoplay, Disney+ and HBO Max ad-supported versions, "
                     "YouTube Premium Lite, Apple Music) is an OPTIONAL PAID add-on exclusive to Vivo Total: R$ 55/mo for 6 months then "
                     "R$ 65/mo on most combos; R$ 80/mo then R$ 95/mo on Total Pro; 5-app version R$ 45/mo then R$ 55/mo on V 1 Giga. "
                     "Standalone value stated as R$ 140.49. It is NOT part of the monthly prices in the Vivo sheet."),
    ("Courtesy windows", "Amazon Prime + Apple Music and Gemini AI Plus are free-for-6-or-12-months courtesies with the after-prices "
                         "stated per row; they are not permanent inclusions."),
    ("Naming", "Two combos display the same on-site name \"Vivo Família 2\"; the \"(+ TV online)\" / \"(+ TV)\" suffixes in the sheet "
               "were added for disambiguation (the cards differ only by the bundled TV package)."),
    ("Session-dependent variants", "A same-day capture by the promo-vs-entry study (repo analysis/promo_vs_entry, 2026-07-17) saw the "
               "500 Mbps + 60 GB tier as \"Vivo Total Essencial\" at a limited-time R$ 130 (list \"de 160\") and Ultra at R$ 170 "
               "(\"de 190\"), while this default headless render served \"Vivo Total Pro\" at a flat R$ 160 and Ultra at a flat R$ 170. "
               "Vivo varies card names/promo prices per session; this sheet records the default public SP render."),
    ("", ""),
    ("CAVEATS (Claro)", ""),
    ("Grid depth", "The full combo grid IS public on claro.com.br/multi (no address/CEP gate for the listing; no checkout entered). "
                   "The recon lead's single R$ 159.90 headline combo is the \"Melhor escolha\" mid-tier — the actual entry combo is "
                   "R$ 129.80 (Fibra 350 Mbps + Controle 41GB)."),
    ("Promo window", "The 350 Mbps + Controle 41GB combo is tagged \"Por tempo limitado\" (limited time)."),
    ("Hidden 6th combo", "A sixth combo (Fibra 600 Mbps + Pós 50GB) exists in the page's embedded JSON but is NOT rendered for SP "
                         "visitors — excluded from the table."),
    ("Program-wide perks", "Stated on the page for Claro Multi as a program (not itemized per combo): up to 35% bill discount vs "
                           "contracting separately; everything on a single bill; Claro Clube advantages; extra smartphone discounts. "
                           "Passaporte Américas = roaming in 46+ countries at no extra cost (page wording: \"mais de 46 países\")."),
    ("", ""),
    ("METHOD", ""),
    ("Language rules", "Headers and descriptive text in English; brand/product names kept as-is (TIM Black, Ultrafibra, Vivo Total, "
                       "Passaporte Américas…). \"Mega\" → Mbps, \"Giga\" → Gbps. Prices dot-decimal, stored as NUMBERS "
                       "(the TIM * is applied via the cell's number format)."),
    ("Raw captures", "analysis/convergent_offers/raw/ (HTML/JSON/screenshots) — kept LOCAL for audit, not committed "
                     "(same practice as the tracker's data/raw)."),
]


def _style_table_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="9AA5B1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    band = PatternFill("solid", fgColor=BAND_FILL)

    for col, (name, width) in enumerate(zip(COLUMNS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    for r, (mobile, internet, streaming, benefits, price, star) in enumerate(rows, start=2):
        values = [mobile, internet, streaming, benefits]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = band
        pcell = ws.cell(row=r, column=5, value=price)
        pcell.number_format = PRICE_FMT_STAR if star else PRICE_FMT
        pcell.alignment = Alignment(horizontal="right", vertical="top")
        if r % 2 == 0:
            pcell.fill = band
    ws.freeze_panes = "A2"


def _style_notes_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 130
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.cell(row=1, column=1, value="Convergent offers snapshot — Notes & Sources").font = Font(bold=True, size=13)
    r = 3
    for label, detail in NOTES:
        if label and not detail:  # section header
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = section_font
            cell.fill = section_fill
            ws.cell(row=r, column=2).fill = section_fill
        elif label:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            c = ws.cell(row=r, column=2, value=detail)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1


def main():
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in (("TIM", TIM_ROWS), ("Vivo", VIVO_ROWS), ("Claro", CLARO_ROWS)):
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = TAB_COLORS[name]
        _style_table_sheet(ws, rows)
    ws = wb.create_sheet("Notes & Sources")
    ws.sheet_properties.tabColor = TAB_COLORS["Notes & Sources"]
    _style_notes_sheet(ws)
    wb.save(OUT)
    print(f"Wrote {OUT.name}: TIM {len(TIM_ROWS)} rows, Vivo {len(VIVO_ROWS)} rows, Claro {len(CLARO_ROWS)} rows + Notes & Sources")


if __name__ == "__main__":
    main()
