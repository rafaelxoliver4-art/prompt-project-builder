# -*- coding: utf-8 -*-
"""check.py — verification for promo_vs_entry.xlsx (CODE TASK #25).

Runnable standalone (``python check.py``) or under pytest. Verifies:
  (a) every ratio_pct (annual + events) recomputes from its two inputs;
  (b) sanity bands: entry controle R$35-80, entry pos R$80-200, ratio 40-400%;
  (c) every non-blank data point carries a source_id that resolves in the sources sheet;
  (d) the 2026 entry anchors equal the tracker's audited values (CODE TASKs #22-#24).
Fail loud, never silently accept.
"""
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(__file__).parent / "promo_vs_entry.xlsx"

AUDITED_2026 = {  # tracker snapshot 2026-07-01, audited #22-#24
    ("TIM", "controle"): 58.99,
    ("Vivo", "controle"): 59.00,
    ("Claro", "controle"): 54.90,   # effective (debito); regular 59.90 checked separately
    ("TIM", "pos"): 129.99,         # cheapest BILL-payment plan (JPMorgan rule)
    ("Vivo", "pos"): 150.00,
    ("Claro", "pos"): 124.90,
}
CLARO_2026_CONTROLE_REGULAR = 59.90

BAND_CONTROLE = (35.0, 80.0)
BAND_POS = (80.0, 200.0)
BAND_RATIO = (40.0, 400.0)


def _rows(ws):
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(header, row))


def _load():
    wb = load_workbook(XLSX, data_only=True)
    return {name: list(_rows(wb[name])) for name in ("annual", "events", "entry_prices", "sources")}


def test_ratios_recompute():
    d = _load()
    checked = 0
    for r in d["annual"]:
        expect = round(r["promo_price"] / r["anchor_price"] * 100, 2)
        assert abs(r["ratio_pct"] - expect) < 0.05, f"annual {r['year']} {r['carrier']}: {r['ratio_pct']} != {expect}"
        # the anchor must equal the entry price of the anchor category in the same row's entry columns
        col = {"controle": "entry_controle", "pos": "entry_pos"}[r["anchor_category"]]
        assert abs(r["anchor_price"] - r[col]) < 0.005, f"annual {r['year']} {r['carrier']}: anchor != {col}"
        checked += 1
    entry = {(e["year"], e["carrier"], e["category"]): e["price_brl"]
             for e in d["entry_prices"] if e["price_brl"] is not None}
    for e in d["events"]:
        anchor = entry[(e["year"], e["carrier"], e["anchor_category"])]
        assert abs(e["anchor_price_brl"] - anchor) < 0.005, f"event '{e['campaign']}': anchor mismatch"
        expect = round(e["promo_price_brl"] / anchor * 100, 2)
        assert abs(e["ratio_pct"] - expect) < 0.05, f"event '{e['campaign']}': {e['ratio_pct']} != {expect}"
        checked += 1
    print(f"  (a) ratios recomputed OK: {checked} rows (12 annual + {checked - 12} events)")


def test_sanity_bands():
    d = _load()
    for e in d["entry_prices"]:
        p = e["price_brl"]
        if p is None:
            assert e["notes"], f"blank cell without a gap note: {e}"
            continue
        if e["category"] == "controle":
            lo, hi = BAND_CONTROLE
            assert lo <= p <= hi, f"controle {e['year']} {e['carrier']} = {p} outside {lo}-{hi}"
        elif e["category"] == "pos":
            lo, hi = BAND_POS
            assert lo <= p <= hi, f"pos {e['year']} {e['carrier']} = {p} outside {lo}-{hi}"
    for sheet in ("annual", "events"):
        for r in d[sheet]:
            lo, hi = BAND_RATIO
            assert lo <= r["ratio_pct"] <= hi, f"{sheet} ratio {r['ratio_pct']} outside {lo}-{hi}"
    print("  (b) sanity bands OK (controle 35-80, pos 80-200, ratio 40-400; blanks carry gap notes)")


def test_every_point_sourced():
    d = _load()
    source_ids = {s["source_id"] for s in d["sources"]}
    n = 0
    for e in d["entry_prices"]:
        if e["price_brl"] is None:
            continue
        assert e["source_id"] in source_ids, f"entry {e['year']} {e['carrier']} {e['category']}: source {e['source_id']} missing"
        n += 1
    for e in d["events"]:
        for sid in str(e["source_ids"]).split(","):
            assert sid.strip() in source_ids, f"event '{e['campaign']}': source {sid} missing"
        n += 1
    for r in d["annual"]:
        for sid in str(r["source_ids"]).split(","):
            assert sid.strip() in source_ids, f"annual {r['year']} {r['carrier']}: source {sid} missing"
        n += 1
    print(f"  (c) source-per-point OK: {n} non-blank data rows, all source_ids resolve ({len(source_ids)} sources)")


def test_2026_anchors_match_tracker():
    d = _load()
    entry = {(e["carrier"], e["category"]): e for e in d["entry_prices"] if e["year"] == 2026}
    for (car, cat), expected in AUDITED_2026.items():
        got = entry[(car, cat)]["price_brl"]
        assert abs(got - expected) < 0.005, f"2026 {car} {cat}: {got} != audited {expected}"
    reg = entry[("Claro", "controle")]["regular_price_brl"]
    assert abs(reg - CLARO_2026_CONTROLE_REGULAR) < 0.005, f"Claro 2026 controle regular {reg} != {CLARO_2026_CONTROLE_REGULAR}"
    print("  (d) 2026 anchors == tracker audited values (incl. Claro 54.90 eff / 59.90 reg; TIM pos 129.99 bill)")


if __name__ == "__main__":
    print(f"check.py — verifying {XLSX.name}")
    test_ratios_recompute()
    test_sanity_bands()
    test_every_point_sourced()
    test_2026_anchors_match_tracker()
    print("ALL GREEN")
