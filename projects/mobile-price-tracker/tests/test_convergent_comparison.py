"""Offline tests for the `convergent_comparison` matrix — CODE TASK #33. No network.

Mirrors the mobile `comparison` discipline: one row per snapshot_date, live exact-date MINIFS over
`convergent_history`, cached values baked to match the formulas exactly, blanks where a carrier sold
no combo of that bundle type, and a build failure that can never break the mobile write.
"""
from datetime import date, datetime

import openpyxl
import pandas as pd

from mobile_tracker.convergent import (BUNDLE_FIBRE_MOBILE, BUNDLE_FIBRE_MOBILE_TV, BUNDLE_FIBRE_TV,
                                       CANONICAL_BUNDLE_TYPES, CONVERGENT_COLUMNS, ConvergentOffer,
                                       bundle_type_of)
from mobile_tracker.excel_writer import (_conv_matrix_value, convergent_bundle_types, write_workbook)
from mobile_tracker.models import Plan

FIRST = 3          # first data row on the matrix (rows 1-2 are the two header rows)


def _plan(price, date_str):
    p = Plan(carrier="tim", category="postpaid", state="SP", plan_name="A", plan_id="tim:1",
             price_brl=price, source_url="x")
    p.snapshot_date, p.snapshot_ts = date_str, f"{date_str}T18:00:00"
    return p


def _co(carrier, name, oid, price, date_str, *, mobile=True, broadband=True, tv=False, landline=False):
    o = ConvergentOffer(carrier=carrier, state="SP", offer_name=name, offer_id=oid, price_brl=price,
                        has_mobile=mobile, has_broadband=broadband, has_tv=tv, has_landline=landline,
                        source_url="x")
    o.snapshot_date, o.snapshot_ts = date_str, f"{date_str}T18:00:00"
    return o


# ---- bundle_type derivation ------------------------------------------------------------------
def test_bundle_type_derived_from_service_flags():
    assert bundle_type_of(True, True, False) == BUNDLE_FIBRE_MOBILE
    assert bundle_type_of(False, True, True) == BUNDLE_FIBRE_TV
    assert bundle_type_of(True, True, True) == BUNDLE_FIBRE_MOBILE_TV
    # a shape outside the canonical three keeps its OWN identity instead of being folded into one
    assert bundle_type_of(True, False, True) == "Mobile + TV"
    assert bundle_type_of(False, True, False, True) == "Fibre + Landline"
    # fewer than two services is not a convergent offer at all
    assert bundle_type_of(True, False, False) is None
    assert bundle_type_of(False, False, False) is None


def test_offer_exposes_bundle_type_in_its_row():
    o = _co("vivo", "Ultra + TV", "vivo:1", 190.0, "2026-08-03", tv=True)
    assert o.bundle_type == BUNDLE_FIBRE_MOBILE_TV
    row = o.as_row()
    assert row["bundle_type"] == BUNDLE_FIBRE_MOBILE_TV
    assert list(row) == CONVERGENT_COLUMNS              # schema order preserved


def test_bundle_type_backfilled_for_rows_written_before_the_column_existed(tmp_path):
    """A pre-#33 convergent_history has no bundle_type. It is a pure function of the service flags,
    which those rows DO carry, so the whole history self-heals on the next write — otherwise the
    matrix would be blank for every past date."""
    from mobile_tracker.excel_writer import _merge_convergent
    old_cols = [c for c in CONVERGENT_COLUMNS if c != "bundle_type"]
    existing = pd.DataFrame([dict(snapshot_date="2026-08-01", carrier="claro", state="SP",
                                  offer_name="Old", offer_id="claro:x", price_brl=129.8,
                                  has_mobile=True, has_broadband=True, has_tv=False,
                                  has_landline=False)], columns=old_cols)
    fresh = pd.DataFrame([_co("tim", "UC7", "tim:1", 89.99, "2026-08-02").as_row()],
                         columns=CONVERGENT_COLUMNS)
    merged = _merge_convergent(existing, fresh)
    assert list(merged.columns) == CONVERGENT_COLUMNS                       # order enforced
    assert set(merged["bundle_type"]) == {BUNDLE_FIBRE_MOBILE}              # old row backfilled


def test_group_list_is_canonical_plus_any_new_shape():
    assert convergent_bundle_types(pd.DataFrame()) == CANONICAL_BUNDLE_TYPES     # stable when empty
    conv = pd.DataFrame({"bundle_type": [BUNDLE_FIBRE_MOBILE, "Mobile + TV", None]})
    assert convergent_bundle_types(conv) == CANONICAL_BUNDLE_TYPES + ["Mobile + TV"]


# ---- the matrix ------------------------------------------------------------------------------
def _wb_with_two_days(tmp_path):
    out = tmp_path / "w.xlsx"
    d1 = [_co("tim", "UC7", "tim:1", 89.99, "2026-08-03"),
          _co("tim", "UC6", "tim:2", 139.99, "2026-08-03"),
          _co("vivo", "Pro", "vivo:1", 160.0, "2026-08-03"),
          _co("vivo", "Ultra+TV", "vivo:2", 190.0, "2026-08-03", tv=True),
          _co("claro", "Multi 350", "claro:1", 129.80, "2026-08-03"),
          _co("claro", "Fibra+TV", "claro:2", 219.80, "2026-08-03", mobile=False, tv=True)]
    d2 = [_co("tim", "UC7", "tim:1", 94.99, "2026-08-04"),
          _co("claro", "Multi 350", "claro:1", 139.80, "2026-08-04")]
    write_workbook([_plan(100.0, "2026-08-03")], out, datetime(2026, 8, 3, 18), convergent=d1)
    write_workbook([_plan(100.0, "2026-08-04")], out, datetime(2026, 8, 4, 18), convergent=d2)
    return out


def test_matrix_layout_rows_per_date_earliest_first(tmp_path):
    ws = openpyxl.load_workbook(_wb_with_two_days(tmp_path), data_only=True)["convergent_comparison"]
    assert ws.cell(1, 1).value == "Date"
    groups = [ws.cell(1, 2 + g * 3).value for g in range(len(CANONICAL_BUNDLE_TYPES))]
    assert groups == CANONICAL_BUNDLE_TYPES
    assert [ws.cell(2, c).value for c in (2, 3, 4)] == ["TIM", "Vivo", "Claro"]
    rows = [r for r in range(FIRST, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, date)]
    assert len(rows) == 2                                            # one row per snapshot_date
    dates = [ws.cell(r, 1).value.date().isoformat() for r in rows]
    assert dates == ["2026-08-03", "2026-08-04"]                     # EARLIEST first (auto-growing)


def test_cells_are_exact_date_minifs_over_convergent_history(tmp_path):
    ws = openpyxl.load_workbook(_wb_with_two_days(tmp_path), data_only=False)["convergent_comparison"]
    f = ws.cell(FIRST, 2).value
    assert f.startswith("=IFERROR(") and "_xlfn.MINIFS(convergent_history!" in f
    assert '"tim"' in f and f'"{BUNDLE_FIBRE_MOBILE}"' in f
    # the locale-safe TEXT date (convergent_history stores snapshot_date as text, like `history`)
    assert f'YEAR($A{FIRST})&"-"&TEXT(MONTH($A{FIRST}),"00")&"-"&TEXT(DAY($A{FIRST}),"00")' in f
    assert 'IF(' in f and '=0,""' in f                               # 0 / no match → blank


def test_cheapest_within_bundle_type_and_blanks_where_absent(tmp_path):
    ws = openpyxl.load_workbook(_wb_with_two_days(tmp_path), data_only=True)["convergent_comparison"]
    # Fibre + Mobile block (cols B/C/D) on the first date
    assert ws.cell(FIRST, 2).value == 89.99          # TIM: cheapest of 89,99 / 139,99
    assert ws.cell(FIRST, 3).value == 160.0          # Vivo
    assert ws.cell(FIRST, 4).value == 129.80         # Claro
    # Fibre + TV block (E/F/G): only Claro sells one → the others are BLANK, never 0
    assert ws.cell(FIRST, 5).value is None and ws.cell(FIRST, 6).value is None
    assert ws.cell(FIRST, 7).value == 219.80
    # Fibre + Mobile + TV block (H/I/J): only Vivo
    assert ws.cell(FIRST, 8).value is None
    assert ws.cell(FIRST, 9).value == 190.0
    assert ws.cell(FIRST, 10).value is None
    # the TV-bearing Vivo offer must NOT leak into the Fibre+Mobile column
    assert ws.cell(FIRST, 3).value == 160.0


def test_baked_values_match_the_formula_helper(tmp_path):
    """The bake must reproduce exactly what MINIFS computes (#21 discipline)."""
    out = _wb_with_two_days(tmp_path)
    conv = pd.read_excel(out, sheet_name="convergent_history", dtype={"snapshot_date": str})
    ws = openpyxl.load_workbook(out, data_only=True)["convergent_comparison"]
    groups = convergent_bundle_types(conv)
    rows = [r for r in range(FIRST, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, date)]
    checked = 0
    for r in rows:
        day = ws.cell(r, 1).value.date().isoformat()
        for g, bundle in enumerate(groups):
            for k, carrier in enumerate(("tim", "vivo", "claro")):
                expected = _conv_matrix_value(conv, carrier, bundle, day)
                assert ws.cell(r, 2 + g * 3 + k).value == expected, (day, bundle, carrier)
                checked += 1
    assert checked == len(rows) * len(groups) * 3


def test_matrix_has_a_colour_scale_per_bundle_block_and_a_note(tmp_path):
    out = _wb_with_two_days(tmp_path)
    ws = openpyxl.load_workbook(out)["convergent_comparison"]
    assert len(list(ws.conditional_formatting)) >= len(CANONICAL_BUNDLE_TYPES)
    assert ws.freeze_panes == f"B{FIRST}"
    assert ws.sheet_view.showGridLines is False
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    note = next((t for t in texts if "BLANK cell" in t), None)
    assert note and "bundle type" in note.lower()


def test_empty_convergent_history_still_yields_a_header_only_sheet(tmp_path):
    out = tmp_path / "w.xlsx"
    write_workbook([_plan(100.0, "2026-08-03")], out, datetime(2026, 8, 3, 18))   # no convergent
    ws = openpyxl.load_workbook(out)["convergent_comparison"]
    assert ws.cell(1, 1).value == "Date"
    assert [ws.cell(1, 2 + g * 3).value for g in range(3)] == CANONICAL_BUNDLE_TYPES
    assert not [r for r in range(FIRST, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, date)]


def test_comparison_sheet_failure_never_breaks_the_mobile_write(tmp_path, monkeypatch, capsys):
    """Guard: the matrix is a side-domain view. If building it raises, the mobile workbook (and the
    collected convergent history) must still be written correctly — and the fail-loud bake check must
    not abort the run over the skipped sheet's values."""
    import mobile_tracker.excel_writer as ew

    def boom(*a, **k):
        raise RuntimeError("matrix exploded")

    monkeypatch.setattr(ew, "_write_convergent_comparison", boom)
    out = tmp_path / "w.xlsx"
    res = ew.write_workbook([_plan(119.99, "2026-08-03")], out, datetime(2026, 8, 3, 18),
                            convergent=[_co("tim", "UC7", "tim:1", 89.99, "2026-08-03")])
    assert "comparison sheet failed" in capsys.readouterr().out
    assert res["plans_in_latest"] == 1 and res["convergent_rows"] == 1     # both domains still written
    wb = openpyxl.load_workbook(out, data_only=True)
    assert "convergent_history" in wb.sheetnames and "comparison" in wb.sheetnames
    assert wb["comparison"].cell(FIRST, 5).value == 119.99                # mobile matrix still baked
