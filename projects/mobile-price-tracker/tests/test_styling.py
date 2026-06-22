"""Offline tests for the visual house style (CODE TASK #11).

Build a small workbook and assert the design-system attributes: gridlines off on every sheet,
tab colors set, conditional-formatting rules on the comparison + carrier sheets, banded data rows.
Data integrity is covered by the other test modules; this only checks presentation.
"""
from datetime import datetime

import openpyxl
import pytest

from mobile_tracker.excel_writer import write_workbook
from mobile_tracker.models import Plan

_ROWS = [
    ("vivo", "postpaid", "Vivo Pós Amazon", 150.0, None, 60.0),
    ("vivo", "postpaid", "Vivo Pós Netflix", 180.0, None, 60.0),
    ("vivo", "lite", "Easy Lite", 30.0, None, 30.0),
    ("claro", "postpaid", "Pós 60GB", 124.9, None, 60.0),
    ("claro", "control", "Controle 25GB", 59.9, 54.9, 25.0),   # has a promo
    ("tim", "postpaid", "TIM Black 70GB", 129.99, None, 70.0),
    ("tim", "control", "TIM Controle 41GB", 58.99, None, 41.0),
    ("tim", "prepaid", "TIM Pré 6GB", 20.0, None, 6.0),
]


@pytest.fixture
def wb(tmp_path):
    plans = []
    for c, cat, name, price, promo, gb in _ROWS:
        p = Plan(carrier=c, category=cat, state="SP", plan_name=name, plan_id=f"{c}:{name}",
                 price_brl=price, price_promo_brl=promo, data_gb=gb, source_url="x")
        p.snapshot_date = "2026-06-22"
        p.snapshot_ts = "2026-06-22T18:00:00"
        plans.append(p)
    out = tmp_path / "styled.xlsx"
    write_workbook(plans, out, datetime(2026, 6, 22, 18))
    return openpyxl.load_workbook(out)


def test_gridlines_off_on_every_sheet(wb):
    for name in wb.sheetnames:
        assert wb[name].sheet_view.showGridLines is False, name


def test_tab_colors_set(wb):
    expected = {"Vivo": "660099", "Claro": "DA291C", "TIM": "0033A0",
                "comparison": "2E7D32", "summary": "102A43"}
    for name, rgb in expected.items():
        tc = wb[name].sheet_properties.tabColor
        assert tc is not None and tc.rgb.endswith(rgb), name


def test_conditional_formatting_on_comparison_and_carrier_sheets(wb):
    assert len(list(wb["comparison"].conditional_formatting)) >= 1
    for carrier in ("Vivo", "Claro", "TIM"):
        assert len(list(wb[carrier].conditional_formatting)) >= 1, carrier
    assert len(list(wb["latest"].conditional_formatting)) >= 1   # price scale + promo accent


def test_data_rows_are_banded(wb):
    ws = wb["latest"]
    banded = any(
        (ws.cell(row=r, column=1).fill is not None
         and ws.cell(row=r, column=1).fill.fgColor is not None
         and str(ws.cell(row=r, column=1).fill.fgColor.rgb).endswith("F2F5FA"))
        for r in range(2, ws.max_row + 1)
    )
    assert banded, "expected at least one banded data row in `latest`"


def test_carrier_sheet_has_subheader_accent(wb):
    ws = wb["Vivo"]
    # the first category sub-header row (row 2) should carry the #D6E0F0 accent fill
    cell = ws.cell(row=2, column=1)
    assert str(cell.fill.fgColor.rgb).endswith("D6E0F0")
