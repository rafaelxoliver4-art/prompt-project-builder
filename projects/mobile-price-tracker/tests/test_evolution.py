"""Offline tests for the DAILY price-evolution view (#12/#13/#16/#17/#18/#19).

#19 changed the matrix from MONTHLY rows back to DAILY rows: `comparison` has one row per distinct
snapshot_date (earliest first), each value cell a live EXACT-date MINIFS over `history` (a locale-safe
text date built from the row's date — history stores dates as TEXT, see excel_writer/_minifs_day). The
table-only layout (#17) is unchanged: group titles row 1, TIM/Vivo/Claro row 2, day rows from row 3,
frozen at B3, not protected, the active sheet; the 4 charts live on the `Charts` sheet via cross-sheet
refs. The heatmap is now a 2-color green→yellow scale (JPMorgan style), per carrier column, no red.
"""
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime

import openpyxl
import pandas as pd
import pytest

from mobile_tracker.excel_writer import write_workbook, evolution_dates, evolution_months
from mobile_tracker.models import Plan

# Table-only `comparison` (#17): header rows 1/2, day data from row 3 (no chart offset).
H1, H2, FIRST = 1, 2, 3


def _sheet_xml_path(z: zipfile.ZipFile, sheet_name: str) -> str:
    """Map a sheet display name → its xl/worksheets/sheetN.xml path (via workbook.xml + rels)."""
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sh in wb.find("m:sheets", ns):
        if sh.get("name") == sheet_name:
            rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rid_to_target[rid].lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise AssertionError(f"sheet {sheet_name!r} not found")


def _mk(carrier, cat, name, price, date_str):
    p = Plan(carrier=carrier, category=cat, state="SP", plan_name=name,
             plan_id=f"{carrier}:{name}", price_brl=price, source_url="x")
    p.snapshot_date = date_str
    p.snapshot_ts = f"{date_str}T18:00:00"
    return p


def _multi_day(path):
    """History with 3 distinct snapshot_dates → the matrix should have one row per date, earliest first."""
    for d, tim, vivo, claro in [("2026-06-22", 122.0, 150.0, 125.0),
                                ("2026-06-23", 121.0, 150.0, 124.9),
                                ("2026-06-24", 119.0, 149.0, 124.0)]:
        plans = [_mk("tim", "postpaid", "TB", tim, d),
                 _mk("vivo", "postpaid", "VP", vivo, d),
                 _mk("claro", "postpaid", "CP", claro, d),
                 _mk("vivo", "lite", "EL", 30.0, d),
                 _mk("claro", "flex", "FX", 45.0, d)]
        write_workbook(plans, path, datetime(2026, 6, int(d[-2:]), 18))
    return path


def test_matrix_rows_are_days(tmp_path):
    wb = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))
    ws = wb["comparison"]
    assert ws.cell(row=H1, column=1).value == "Date"
    titles = [ws.cell(row=H1, column=c).value for c in (2, 5, 8, 11)]
    assert titles == ["Control (R$/mo)", "Post (R$/mo)", "Pre (R$/mo, 30-day)", "Digital (R$/mo)"]
    assert [ws.cell(row=H2, column=c).value for c in (2, 3, 4)] == ["TIM", "Vivo", "Claro"]
    # one row per snapshot_date, stored as real dates, EARLIEST first
    d1 = ws.cell(row=FIRST, column=1).value
    d2 = ws.cell(row=FIRST + 1, column=1).value
    d3 = ws.cell(row=FIRST + 2, column=1).value
    assert (d1.year, d1.month, d1.day) == (2026, 6, 22)       # 2026-06-22 is the FIRST row
    assert (d2.year, d2.month, d2.day) == (2026, 6, 23)
    assert (d3.year, d3.month, d3.day) == (2026, 6, 24)
    assert ws.cell(row=FIRST + 3, column=1).value is None       # only three dates present


def test_day_cells_display_dd_mmm(tmp_path):
    ws = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))["comparison"]
    assert ws.cell(row=FIRST, column=1).number_format == "dd-mmm"      # 2026-06-22 → "22-Jun"
    assert ws.cell(row=FIRST + 2, column=1).number_format == "dd-mmm"


def test_value_cells_are_exact_date_minifs_over_history(tmp_path):
    ws = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))["comparison"]
    post_tim = ws.cell(row=FIRST, column=5).value          # Post group, TIM column
    assert isinstance(post_tim, str)
    assert "MINIFS" in post_tim and "history!" in post_tim
    assert '"tim"' in post_tim and '"postpaid"' in post_tim
    # EXACT date = a text date built from the date cell ($A3): YEAR/MONTH/DAY, NO wildcard
    assert f"YEAR($A{FIRST})" in post_tim and f"DAY($A{FIRST})" in post_tim
    assert '"-*"' not in post_tim                            # exact date, NOT a month prefix
    # digital = cheapest of lite OR flex, same exact-date scope, via the reciprocal-max min trick
    digital_vivo = ws.cell(row=FIRST, column=12).value
    assert '"lite"' in digital_vivo and '"flex"' in digital_vivo and f"DAY($A{FIRST})" in digital_vivo
    assert "MAX(" in digital_vivo and "1/" in digital_vivo


def test_distinct_dates_one_row_each(tmp_path):
    """Daily: two snapshots on DIFFERENT dates → TWO rows (one per date), not a roll-up; history keeps both."""
    out = tmp_path / "m.xlsx"
    write_workbook([_mk("tim", "postpaid", "TB", 122.0, "2026-06-22")], out, datetime(2026, 6, 22, 18))
    write_workbook([_mk("tim", "postpaid", "TB", 118.0, "2026-06-23")], out, datetime(2026, 6, 23, 18))
    ws = openpyxl.load_workbook(out)["comparison"]
    d1 = ws.cell(row=FIRST, column=1).value
    d2 = ws.cell(row=FIRST + 1, column=1).value
    assert (d1.month, d1.day) == (6, 22) and (d2.month, d2.day) == (6, 23)   # two distinct daily rows
    assert ws.cell(row=FIRST + 2, column=1).value is None
    hist = pd.read_excel(out, sheet_name="history")
    assert set(hist["snapshot_date"].astype(str)) == {"2026-06-22", "2026-06-23"}


def _mk_pre(carrier, name, price, date_str, validity):
    p = _mk(carrier, "prepaid", name, price, date_str)
    p.validity_days = validity
    return p


def test_pre_column_filters_30day_validity(tmp_path):
    """#18: the Pre cell is a MINIFS that ALSO filters validity_days >= 28, so it picks the cheapest
    30-DAY prepaid plan on that date — not a cheaper short-validity tier (the old R$1/day-style bug)."""
    from openpyxl.utils import get_column_letter
    from mobile_tracker.models import COLUMNS
    out = tmp_path / "m.xlsx"
    # Claro: a cheaper 15-day tier AND the real 30-day tier on the same date → Pre must pick the 30-day.
    write_workbook([
        _mk_pre("claro", "Prezão 5GB (15d)", 15.0, "2026-06-22", 15),
        _mk_pre("claro", "Prezão 12GB (30d)", 30.0, "2026-06-22", 30),
    ], out, datetime(2026, 6, 22, 18))
    wb = openpyxl.load_workbook(out)
    assert "validity_days" in [c.value for c in wb["history"][1]]    # history carries the column
    ws = wb["comparison"]
    pre_claro = ws.cell(row=FIRST, column=10).value     # Pre group cols 8/9/10 = TIM/Vivo/Claro
    assert isinstance(pre_claro, str) and "MINIFS" in pre_claro and '"prepaid"' in pre_claro
    vcol = get_column_letter(COLUMNS.index("validity_days") + 1)
    assert f'history!${vcol}:${vcol},">=28"' in pre_claro    # the 30-day validity filter


def _mk_post(carrier, name, price, date_str, payment_method=None):
    p = _mk(carrier, "postpaid", name, price, date_str)
    p.payment_method = payment_method
    return p


def test_post_column_tracks_cheapest_including_credit_card(tmp_path):
    """#27 (Bridge reverted #24's exclusion): the matrix Post pick is the plain CHEAPEST postpaid plan —
    TIM's R$119,99 credit-card "no cartão" plan IS the tracked entry-level (a change on it is what we
    want to detect). The payment_method TAG is still recorded in history (billing-type context kept)."""
    out = tmp_path / "m.xlsx"
    write_workbook([
        _mk_post("tim", "TIM Black A Express 67GB", 119.99, "2026-07-01", "credit_card"),
        _mk_post("tim", "TIM Black 70GB", 129.99, "2026-07-01", "bill"),
        _mk_post("vivo", "Vivo Pós", 150.0, "2026-07-01", None),
    ], out, datetime(2026, 7, 1, 18))

    # both TIM postpaid plans retained, still tagged with their billing type
    hist = pd.read_excel(out, sheet_name="history")
    assert "payment_method" in hist.columns
    tim_post = hist[(hist["carrier"] == "tim") & (hist["category"] == "postpaid")]
    assert {round(float(x), 2) for x in tim_post["price_brl"]} == {119.99, 129.99}
    assert set(tim_post["payment_method"]) == {"credit_card", "bill"}

    fwb = openpyxl.load_workbook(out, data_only=False)
    vwb = openpyxl.load_workbook(out, data_only=True)
    post_tim_f = fwb["comparison"].cell(FIRST, 5).value               # Post group cols 5/6/7 = TIM/Vivo/Claro
    assert '"<>credit_card"' not in post_tim_f                        # no exclusion anywhere in the pick
    assert float(vwb["comparison"].cell(FIRST, 5).value) == 119.99   # baked = the CHEAPEST (credit-card) plan
    assert float(vwb["comparison"].cell(FIRST, 6).value) == 150.0


def test_matrix_value_postpaid_plain_cheapest():
    """#27 unit: _matrix_value postpaid = plain min over the category — credit_card included, so bake ==
    the formula (which carries no payment criterion)."""
    from mobile_tracker.excel_writer import _matrix_value
    tagged = pd.DataFrame([
        {"carrier": "tim", "category": "postpaid", "snapshot_date": "2026-07-01",
         "price_brl": 119.99, "payment_method": "credit_card"},
        {"carrier": "tim", "category": "postpaid", "snapshot_date": "2026-07-01",
         "price_brl": 129.99, "payment_method": "bill"},
    ])
    assert _matrix_value(tagged, "tim", "postpaid", "single", "2026-07-01") == 119.99  # cheapest, incl. credit


def test_digital_lite_requires_sem_fidelidade(tmp_path):
    """#26: the Digital Vivo/lite pick uses the cheapest plan that OFFERS a 'Sem fidelidade' choice
    (loyalty_months set — a monthly/annual toggle card). A monthly-ONLY tier (20GB 'Plano mensal' R$35,
    no loyalty) is EXCLUDED, so the entry-level is the cheapest Sem-fidelidade plan (R$45) — not R$35.
    Flex (Claro) has no toggle → unrestricted. All lite plans stay in history."""
    from openpyxl.utils import get_column_letter
    from mobile_tracker.models import COLUMNS
    out = tmp_path / "m.xlsx"
    el20 = _mk("vivo", "lite", "Easy Lite 20", 35.0, "2026-07-01")        # monthly-only, NO loyalty
    el30 = _mk("vivo", "lite", "Easy Lite 30", 45.0, "2026-07-01"); el30.loyalty_months = 12
    el40 = _mk("vivo", "lite", "Easy Lite 40", 55.0, "2026-07-01"); el40.loyalty_months = 12
    flex = _mk("claro", "flex", "Flex 15", 44.9, "2026-07-01")            # no toggle → unrestricted
    write_workbook([el20, el30, el40, flex], out, datetime(2026, 7, 1, 18))
    fwb = openpyxl.load_workbook(out, data_only=False)
    vwb = openpyxl.load_workbook(out, data_only=True)
    lc = get_column_letter(COLUMNS.index("loyalty_months") + 1)
    dig_vivo_f = fwb["comparison"].cell(FIRST, 12).value                  # Digital group K/L/M = TIM/Vivo/Claro
    assert f'history!${lc}:${lc},">0"' in dig_vivo_f                      # lite gated on loyalty_months
    assert float(vwb["comparison"].cell(FIRST, 12).value) == 45.0        # cheapest Sem-fidelidade, NOT the R$35
    assert float(vwb["comparison"].cell(FIRST, 13).value) == 44.9        # Claro flex unchanged
    hist = pd.read_excel(out, sheet_name="history")
    lite_prices = {round(float(x), 2) for x in hist[(hist.carrier == "vivo") & (hist.category == "lite")]["price_brl"]}
    assert {35.0, 45.0, 55.0}.issubset(lite_prices)                       # all 3 kept in history (incl. the 20GB)


def test_matrix_value_digital_lite_loyalty_gated():
    """#26 unit: _matrix_value digital keeps only lite plans with loyalty_months>0; blanks (pre-#23 lite
    with no loyalty) self-heal to None; flex is unrestricted."""
    from mobile_tracker.excel_writer import _matrix_value
    df = pd.DataFrame([
        {"carrier": "vivo", "category": "lite", "snapshot_date": "2026-07-01", "price_brl": 35.0, "loyalty_months": None},
        {"carrier": "vivo", "category": "lite", "snapshot_date": "2026-07-01", "price_brl": 45.0, "loyalty_months": 12},
    ])
    assert _matrix_value(df, "vivo", None, "digital", "2026-07-01") == 45.0     # R$35 (no loyalty) excluded
    old = pd.DataFrame([
        {"carrier": "vivo", "category": "lite", "snapshot_date": "2026-06-22", "price_brl": 30.0, "loyalty_months": None}])
    assert _matrix_value(old, "vivo", None, "digital", "2026-06-22") is None    # pre-#23 → blank (self-heal)


def test_digital_tim_fit_mensal_sets_entry(tmp_path):
    """#28: TIM Controle Fit joins the Digital column. Its two versions are SEPARATE plans — the Anual
    (R$30, loyalty_months=12, 12-mo permanence) must NOT set the entry; the no-commitment Mensal (R$35,
    loyalty blank) does. The formula gains a fit MINIFS gated on loyalty_months BLANK (criterion "=");
    lite/flex behavior unchanged; both fit versions stay in history."""
    from openpyxl.utils import get_column_letter
    from mobile_tracker.models import COLUMNS
    out = tmp_path / "m.xlsx"
    anual = _mk("tim", "fit", "TIM Controle Fit Anual", 30.0, "2026-07-13"); anual.loyalty_months = 12
    mensal = _mk("tim", "fit", "TIM Controle Fit Mensal", 35.0, "2026-07-13")     # no commitment
    el30 = _mk("vivo", "lite", "Easy Lite 30", 45.0, "2026-07-13"); el30.loyalty_months = 12
    flex = _mk("claro", "flex", "Flex 15", 44.9, "2026-07-13")
    write_workbook([anual, mensal, el30, flex], out, datetime(2026, 7, 13, 18))
    fwb = openpyxl.load_workbook(out, data_only=False)
    vwb = openpyxl.load_workbook(out, data_only=True)
    lc = get_column_letter(COLUMNS.index("loyalty_months") + 1)
    dig_tim_f = fwb["comparison"].cell(FIRST, 11).value                   # Digital group K/L/M = TIM/Vivo/Claro
    assert '"fit"' in dig_tim_f                                           # fit joined the Digital pick
    assert f'history!${lc}:${lc},"="' in dig_tim_f                        # …gated on loyalty BLANK
    assert float(vwb["comparison"].cell(FIRST, 11).value) == 35.0        # TIM Digital = the Mensal, not R$30
    assert float(vwb["comparison"].cell(FIRST, 12).value) == 45.0        # Vivo lite unchanged (#26)
    assert float(vwb["comparison"].cell(FIRST, 13).value) == 44.9        # Claro flex unchanged
    hist = pd.read_excel(out, sheet_name="history")
    fitp = {round(float(x), 2) for x in hist[(hist.carrier == "tim") & (hist.category == "fit")]["price_brl"]}
    assert fitp == {30.0, 35.0}                                           # both versions kept in history


def test_matrix_value_digital_fit_requires_no_commitment():
    """#28 unit: _matrix_value digital keeps only fit plans with loyalty_months BLANK (the Mensal);
    the Anual (loyalty 12) is excluded — mirroring the formula's "=" criterion so bake == formula."""
    from mobile_tracker.excel_writer import _matrix_value
    df = pd.DataFrame([
        {"carrier": "tim", "category": "fit", "snapshot_date": "2026-07-13", "price_brl": 30.0, "loyalty_months": 12},
        {"carrier": "tim", "category": "fit", "snapshot_date": "2026-07-13", "price_brl": 35.0, "loyalty_months": None},
    ])
    assert _matrix_value(df, "tim", None, "digital", "2026-07-13") == 35.0      # Mensal; Anual excluded
    only_anual = df[df["loyalty_months"].notna()]
    assert _matrix_value(only_anual, "tim", None, "digital", "2026-07-13") is None   # loyalty-only → blank


def test_heatmap_colorscale_present(tmp_path):
    wb = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))
    assert len(list(wb["comparison"].conditional_formatting)) >= 1


def test_heatmap_is_per_block_green_yellow_no_red(tmp_path):
    """#20: comparison heatmap = ONE 2-colour green→yellow scale per category BLOCK (its 3 carrier
    columns × all date rows), JPMorgan style, NO red. _multi_day has 3 dates → rows 3..5."""
    ws = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))["comparison"]
    blocks = {}
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            if rule.type == "colorScale" and rule.colorScale is not None:
                blocks[str(cf.sqref)] = [str(c.rgb) for c in rule.colorScale.color]
    # exactly the 4 per-category blocks (Control B:D, Post E:G, Pre H:J, Digital K:M), over the day rows
    assert set(blocks) == {"B3:D5", "E3:G5", "H3:J5", "K3:M5"}
    for colors in blocks.values():
        assert len(colors) == 2                              # 2-colour (not the 3-colour red scale)
        assert any("A9D08E" in c for c in colors) and any("FFE699" in c for c in colors)
    flat = "".join(c for cols in blocks.values() for c in cols)
    assert "F8696B" not in flat                              # no red


def _color_scales(ws):
    out = []
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            if rule.type == "colorScale" and rule.colorScale is not None:
                out.append([str(c.rgb) for c in rule.colorScale.color])
    return out


def test_other_sheets_keep_3color_red_scale(tmp_path):
    """#19: ONLY `comparison` switched to green→yellow; latest / Ranking keep the 3-colour RED scale."""
    wb = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))
    for name in ("latest", "Ranking"):
        scales = _color_scales(wb[name])
        assert scales, f"{name} has no color-scale rule"
        flat = "".join(c for s in scales for c in s)
        assert "F8696B" in flat                             # the red end colour is retained
        assert any(len(s) == 3 for s in scales)             # still a 3-colour scale (not the 2-colour one)


def test_four_line_charts_reference_daily_matrix(tmp_path):
    out = _multi_day(tmp_path / "m.xlsx")
    with zipfile.ZipFile(out) as z:
        all_charts = [z.read(n).decode("utf-8")
                      for n in z.namelist() if re.match(r"xl/charts/chart\d+\.xml$", n)]
    line_xmls = [x for x in all_charts if "lineChart" in x]     # exclude the #20 bar chart
    assert len(line_xmls) == 4                                  # Control / Post / Pre / Digital
    for xml in line_xmls:
        assert "'comparison'!$A" in xml                         # X = the Date column
        series_vals = re.findall(r"'comparison'!\$[B-M]\$\d+:\$[B-M]\$\d+", xml)
        assert len(series_vals) >= 3                            # TIM/Vivo/Claro value ranges (data rows)
        # series NAMES resolve to the row-2 carrier header (titles_from_data) → legend shows TIM/Vivo/Claro
        name_refs = re.findall(r"'comparison'!\$?[B-M]\$?2(?![0-9])", xml)
        assert len(name_refs) >= 3
    joined = "".join(line_xmls)
    for col in "BCDEFGHIJKLM":                                  # all 12 carrier×category columns charted
        assert f"'comparison'!${col}$" in joined
    for color in ("0033A0", "660099", "DA291C"):               # palette line colors (TIM/Vivo/Claro)
        assert color in joined


def test_current_price_bar_chart_on_charts(tmp_path):
    """#20: a current-price grouped BAR chart on `Charts`, reading the EXACT LAST matrix row via a helper
    table of cross-sheet refs (structure-driven, auto-updating). 4 categories × 3 carrier series."""
    out = _multi_day(tmp_path / "m.xlsx")
    wb = openpyxl.load_workbook(out)
    comp = wb["comparison"]
    # the true last data row: col A holds dates (the note below it is a string)
    last = max(r for r in range(FIRST, comp.max_row + 1)
               if isinstance(comp.cell(r, 1).value, date))
    refs = [c.value for row in wb["Charts"].iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=") and "comparison" in c.value]
    assert len(refs) == 12                                      # 4 categories × 3 carriers
    # every helper ref points at the EXACT matrix last row (no-offer cells wrapped in IF(...="",NA(),...))
    for r in refs:
        rows = {int(m) for m in re.findall(r"'comparison'!\$?[A-M]\$?(\d+)", r)}
        assert rows == {last}
    with zipfile.ZipFile(out) as z:
        cx = [z.read(n).decode("utf-8") for n in z.namelist() if re.match(r"xl/charts/chart\d+\.xml$", n)]
    assert sum("barChart" in x for x in cx) == 1                # exactly one grouped bar chart
    assert sum("lineChart" in x for x in cx) == 4               # plus the 4 line charts (5 total)
    bar = next(x for x in cx if "barChart" in x)
    assert "'Charts'!" in bar                                   # bar reads the on-sheet helper table


def _chart_xmls(path):
    with zipfile.ZipFile(path) as z:
        return [z.read(n).decode("utf-8") for n in z.namelist()
                if re.match(r"xl/charts/chart\d+\.xml$", n)]


def _axis_block(xml, axis):
    m = re.search(rf"<(?:c:)?{axis}>(.*?)</(?:c:)?{axis}>", xml, re.S)
    return m.group(1) if m else ""


def test_line_charts_jpmorgan_style(tmp_path):
    """#25: the 4 line charts match JPMorgan Figs 7/8/9 — title '<Cat> – Monthly Prices', a FIXED Y range
    per category (stable view as history grows), bottom legend, carrier palette, NO vertical gridlines and
    ONE faint horizontal gridline. Automated: still cross-sheet refs into the matrix (checked elsewhere)."""
    from mobile_tracker.excel_writer import EVOLUTION_YRANGE
    out = _multi_day(tmp_path / "m.xlsx")
    lines = [x for x in _chart_xmls(out) if "lineChart" in x]
    assert len(lines) == 4
    seen = {}
    for xml in lines:
        tm = re.search(r"<a:t>([^<]*Monthly Prices)</a:t>", xml)
        assert tm, "line chart title must be '<Cat> – Monthly Prices'"
        short = tm.group(1).split()[0]                          # Control / Post / Pre / Digital
        val, cat = _axis_block(xml, "valAx"), _axis_block(xml, "catAx")
        ymin = re.search(r'<(?:c:)?min val="([\d.]+)"', val)
        ymax = re.search(r'<(?:c:)?max val="([\d.]+)"', val)
        assert ymin and ymax, f"{short}: a FIXED Y range must be set"
        seen[short] = (float(ymin.group(1)), float(ymax.group(1)))
        # legend at bottom, carrier palette present
        assert re.search(r"<(?:c:)?legendPos val=\"b\"", xml)
        assert "0033A0" in xml and "660099" in xml and "DA291C" in xml
        # NO vertical gridlines (category/Date axis), exactly ONE faint horizontal (value axis)
        assert "majorGridlines" not in cat, f"{short}: no vertical gridlines"
        assert "majorGridlines" in val and "ECECEC" in val, f"{short}: one faint horizontal gridline"
    assert seen == {k: (float(v[0]), float(v[1])) for k, v in EVOLUTION_YRANGE.items()}
    # sanity on the actual numbers (Post 90–170, Pre the 30-day MONTHLY value ~20–40, NOT R$/day)
    assert seen["Post"] == (90.0, 170.0) and seen["Pre"] == (20.0, 40.0)


def test_bar_chart_clean_no_heavy_gridlines(tmp_path):
    """#25: the current-price bar chart (Fig-11) is clean — no vertical gridlines, one faint horizontal."""
    out = _multi_day(tmp_path / "m.xlsx")
    bar = next(x for x in _chart_xmls(out) if "barChart" in x)
    assert "majorGridlines" not in _axis_block(bar, "catAx")     # no vertical gridlines
    val = _axis_block(bar, "valAx")
    assert "majorGridlines" in val and "ECECEC" in val           # one faint horizontal
    assert re.search(r"<a:t>Current entry-level price by carrier \(R\$\)</a:t>", bar)


def test_tim_payment_note_on_comparison(tmp_path):
    """#25/#27: the comparison sheet carries the TIM billing-type footnote — the tracked entry-level is
    the cheapest plan (R$119,99, credit-card-only), with the bill-payment R$129,99 noted as context."""
    ws = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))["comparison"]
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    note = next((t for t in texts if "cheapest postpaid plan" in t), None)
    assert note is not None, "TIM payment-method note missing from comparison"
    assert "119,99" in note and "129,99" in note                 # tracked (cheapest) + the bill alternative
    assert "CREDIT-CARD" in note and "payment_method" in note    # billing type flagged + where it lives
    assert "changes" in note                                     # change-detection is the point (#27)


def test_inject_cached_values_handles_all_openpyxl_v_shapes():
    """#21 regression: the injector must handle every way openpyxl serializes a formula cell — empty
    <v/> (3.1.x lxml), empty <v></v> (3.1.x stdlib), and no <v> (older) — producing exactly ONE <v> with
    the value, and be idempotent for a real value. (A guard that only skipped '<v>' silently baked 0 on
    openpyxl 3.1.5.)"""
    from mobile_tracker.excel_writer import _inject_cached_values
    # self-closing empty placeholder (openpyxl 3.1.x + lxml)
    out, n = _inject_cached_values('<s><c r="B3"><f>MINIFS(1)</f><v/></c></s>', {"B3": 58.99})
    assert n == 1 and "<v>58.99</v>" in out and "<v/>" not in out and out.count("<v>") == 1
    # empty <v></v> placeholder (openpyxl 3.1.x + stdlib etree — the reviewer's env)
    out, n = _inject_cached_values('<s><c r="B3"><f>MINIFS(1)</f><v></v></c></s>', {"B3": 150.0})
    assert n == 1 and "<v>150</v>" in out and out.count("<v>") == 1        # 150.0 → "150"
    # no <v> at all (older openpyxl)
    out, n = _inject_cached_values('<s><c r="B3"><f>MINIFS(1)</f></c></s>', {"B3": 44.9})
    assert n == 1 and "<v>44.9</v>" in out
    # idempotent: a real cached value is left unchanged, not doubled
    out, n = _inject_cached_values('<s><c r="B3"><f>F</f><v>99</v></c></s>', {"B3": 1.0})
    assert n == 0 and "<v>99</v>" in out and "<v>1</v>" not in out
    # non-formula cell is skipped; and "B3" must not match "B30"
    out, n = _inject_cached_values('<s><c r="B30"><f>F</f><v/></c></s>', {"B3": 1.0})
    assert n == 0 and "<v>1</v>" not in out


def test_bake_guard_raises_on_incomplete_bake(tmp_path, monkeypatch):
    """#21: if the injector ever fails to bake all cells (e.g. a future openpyxl serialization change
    defeats it), write_workbook must RAISE (fail loud) rather than silently ship a blank workbook."""
    import mobile_tracker.excel_writer as ew
    monkeypatch.setattr(ew, "_inject_cached_values", lambda xml, coords: (xml, 0))   # simulate 0-bake
    with pytest.raises(RuntimeError, match="bake"):
        write_workbook([_mk("tim", "postpaid", "TB", 120.0, "2026-06-22")],
                       tmp_path / "m.xlsx", datetime(2026, 6, 22, 18))


def test_formula_cells_have_cached_values_baked(tmp_path):
    """#21: openpyxl writes formulas WITHOUT cached values → blank in non-recalc viewers. The bake injects
    <v> alongside <f>: data_only=True now shows values AND data_only=False still shows the live formulas."""
    out = _multi_day(tmp_path / "m.xlsx")                 # 3 dates; day 1 (22-Jun) TIM postpaid = 122.0
    fwb = openpyxl.load_workbook(out, data_only=False)    # formulas
    vwb = openpyxl.load_workbook(out, data_only=True)     # cached values
    comp_f, comp_v = fwb["comparison"], vwb["comparison"]
    post_tim_f = comp_f.cell(FIRST, 5).value             # Post group, TIM, first day
    assert isinstance(post_tim_f, str) and "MINIFS" in post_tim_f          # formula still present
    assert comp_v.cell(FIRST, 5).value is not None                        # cached value baked
    assert float(comp_v.cell(FIRST, 5).value) == 122.0                    # ...and correct
    # every matrix cell that has a formula AND a real (non-blank) value should now carry a cached value
    baked = sum(1 for r in range(FIRST, FIRST + 3) for c in range(2, 14)
                if isinstance(comp_f.cell(r, c).value, str) and comp_v.cell(r, c).value is not None)
    assert baked >= 9                                     # postpaid/lite/flex populated across 3 days
    # summary KPI: formula + cached value both present (vivo min over latest = 30.0)
    summ_f, summ_v = fwb["summary"], vwb["summary"]
    assert isinstance(summ_f["C9"].value, str) and "MINIFS" in summ_f["C9"].value
    assert float(summ_v["C9"].value) == 30.0


def test_charts_sheet_exists_and_comparison_is_table_only(tmp_path):
    """#17: the 4 charts hang off the `Charts` sheet; `comparison` is a chart-free table."""
    out = _multi_day(tmp_path / "m.xlsx")
    assert "Charts" in openpyxl.load_workbook(out).sheetnames
    with zipfile.ZipFile(out) as z:
        comp_xml = z.read(_sheet_xml_path(z, "comparison")).decode("utf-8")
        charts_xml = z.read(_sheet_xml_path(z, "Charts")).decode("utf-8")
        assert "<drawing" not in comp_xml          # no chart overlay on the table sheet
        assert "<drawing" in charts_xml            # the charts drawing hangs off the Charts sheet


def test_comparison_is_active_sheet(tmp_path):
    """Opening the workbook lands on the clean `comparison` table, not on `history`."""
    wb = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))
    assert wb.active.title == "comparison"
    selected = [ws.title for ws in wb.worksheets if ws.sheet_view.tabSelected]
    assert selected == ["comparison"]              # the ONLY selected tab → Excel opens on it


def test_comparison_table_only_frozen_b3_and_not_protected(tmp_path):
    """The 'locked' symptom is gone: header frozen at B3, sheet never protected, table starts at the top."""
    ws = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))["comparison"]
    assert ws.freeze_panes == "B3"
    assert ws.protection.sheet is False
    assert ws.cell(row=1, column=1).value == "Date"            # header at row 1
    assert ws.cell(row=FIRST, column=1).value is not None      # data at row 3


def test_ranking_view_preserved(tmp_path):
    wb = openpyxl.load_workbook(_multi_day(tmp_path / "m.xlsx"))
    assert "Ranking" in wb.sheetnames
    assert wb["Ranking"].cell(row=1, column=1).value.startswith("Ranking")


def test_evolution_dates_helper():
    hist = pd.DataFrame({"snapshot_date": ["2026-06-22", "2026-06-21", "2026-06-22", None]})
    assert evolution_dates(hist) == ["2026-06-21", "2026-06-22"]    # distinct, earliest first


def test_evolution_months_helper():
    # evolution_months is retained as a helper (no longer used by the daily matrix).
    hist = pd.DataFrame(
        {"snapshot_date": ["2026-06-22", "2026-05-01", "2026-06-10", "2026-07-30", None]})
    assert evolution_months(hist) == [date(2026, 5, 1), date(2026, 6, 1), date(2026, 7, 1)]


def test_same_day_rerun_replaces_not_unions(tmp_path):
    out = tmp_path / "m.xlsx"
    write_workbook([_mk("tim", "postpaid", "A", 120.0, "2026-06-22"),
                    _mk("vivo", "postpaid", "B", 150.0, "2026-06-22")],
                   out, datetime(2026, 6, 22, 18))
    write_workbook([_mk("tim", "postpaid", "A", 117.0, "2026-06-22")], out, datetime(2026, 6, 22, 19))
    hist = pd.read_excel(out, sheet_name="history")
    rows = hist[hist["snapshot_date"].astype(str) == "2026-06-22"]
    assert len(rows) == 1 and float(rows.iloc[0]["price_brl"]) == 117.0
