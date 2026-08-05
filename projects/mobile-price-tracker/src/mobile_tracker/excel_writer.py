"""Write the tracker workbook: history (append-only), latest, changes, summary.

Design (CONTEXT §7):
- We rebuild the whole file each run from a merged dataframe, so `history` is the
  union of all past snapshots + today's, de-duplicated on (date, carrier, category,
  state, plan_name) keeping the latest write. This makes re-running the same day idempotent.
- `latest` = the most recent snapshot only.
- `changes` = diff of latest vs the previous snapshot date (new / removed / price moves).
- `summary` = run metadata + per-carrier KPIs as Excel formulas (computed on open).
"""
from __future__ import annotations

import os
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .convergent import (CANONICAL_BUNDLE_TYPES, COMPARISON_BUNDLE_TYPES, CONVERGENT_COLUMNS,
                         bundle_type_of, migrate_bundle_type_label)
from .models import COLUMNS

FONT = "Arial"
# Canonical identity = (carrier, state, plan_id). plan_id is the stable per-plan key (CONTEXT §4);
# if a row ever lacks plan_id we fall back to plan_name so nothing crashes.
def _with_key(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "plan_id" not in df.columns:
        df["plan_id"] = pd.NA
    pid = df["plan_id"].astype("string")
    df["_key"] = pid.where(pid.notna() & (pid.str.strip() != ""), df["plan_name"].astype("string"))
    return df


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
CURRENCY_FMT = 'R$ #,##0.00;[RED]-R$ #,##0.00;"-"'

# ---- House style (design system, CONTEXT §7) — defined once, applied everywhere ------------
BAND_FILL = PatternFill("solid", fgColor="F2F5FA")           # alternate data rows
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E0F0")      # carrier-sheet category sub-headers
SUBHEADER_FONT = Font(name=FONT, bold=True, color="1F3864")
PROMO_FILL = PatternFill("solid", fgColor="FFF2CC")          # cells with a promo price
CHEAP_FILL = PatternFill("solid", fgColor="C6EFCE")          # cheapest-of-the-row highlight
CHEAP_FONT = Font(name=FONT, bold=True, color="006100")
HEADER_BORDER = Border(bottom=Side(style="thin", color="FFFFFF"))
RIGHT = Alignment(horizontal="right")
# tab colors per sheet (engine/flat sheets neutral grey; views + carriers branded)
TAB_COLORS = {
    "Vivo": "660099", "Claro": "DA291C", "TIM": "0033A0",
    "comparison": "2E7D32", "Charts": "43A047", "Ranking": "1565C0", "summary": "102A43",
    "history": "8A8A8A", "latest": "8A8A8A", "changes": "8A8A8A",
    "convergent_history": "6A1B9A",     # convergent/combo domain (#31) — distinct from the mobile sheets
    "convergent_comparison": "8E24AA",  # the combo evolution matrix (#33) — same family, lighter
}


def _price_scale() -> ColorScaleRule:
    """3-colour scale: green = cheap (low) → yellow → red = expensive (high).
    Used by latest / operator / Ranking sheets. The comparison matrix uses the 2-colour scale below."""
    return ColorScaleRule(start_type="min", start_color="63BE7B",
                          mid_type="percentile", mid_value=50, mid_color="FFEB84",
                          end_type="max", end_color="F8696B")


def _price_scale_2color() -> ColorScaleRule:
    """2-colour green→yellow scale (JPMorgan style, #19): green = cheaper (min) → yellow = pricier (max),
    no red. Applied PER CARRIER COLUMN on the comparison matrix so each carrier's price evolution reads
    down its own column (each column scaled by its own min/max). Soft, professional shades."""
    return ColorScaleRule(start_type="min", start_color="A9D08E",   # soft green = cheaper
                          end_type="max", end_color="FFE699")        # soft yellow = pricier


def _gridless(ws, tab_color: str | None = None):
    ws.sheet_view.showGridLines = False          # clean modern look — no gridlines anywhere
    if tab_color:
        ws.sheet_properties.tabColor = tab_color


def _house_header(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(vertical="center", horizontal="left")


def _band(ws, first_row: int, last_row: int, n_cols: int):
    """Shade every other data row (first row plain, second shaded, …)."""
    for i, row in enumerate(range(first_row, last_row + 1)):
        if i % 2 == 1:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = BAND_FILL


def _color_scale(ws, col_idx: int, first_row: int, last_row: int):
    if last_row >= first_row:
        L = get_column_letter(col_idx)
        ws.conditional_formatting.add(f"{L}{first_row}:{L}{last_row}", _price_scale())


def _df(plans) -> pd.DataFrame:
    rows = [p.as_row() for p in plans]
    df = pd.DataFrame(rows, columns=COLUMNS)
    return df


def _merge_history(existing: pd.DataFrame, fresh: pd.DataFrame) -> pd.DataFrame:
    # A re-run of a snapshot_date REPLACES that date's rows (idempotent per date): `fresh` fully owns
    # the date(s) it carries, so a same-day re-run / double-fire can't union-inflate a date. Other
    # dates accumulate untouched (append-only across days; CONTEXT §4).
    if not fresh.empty and "snapshot_date" in existing.columns and not existing.empty:
        fresh_dates = set(fresh["snapshot_date"].dropna().unique())
        existing = existing[~existing["snapshot_date"].isin(fresh_dates)]
    df = pd.concat([existing, fresh], ignore_index=True)
    df = _with_key(df)
    # `category` IS part of the identity (#37). Without it a plan_id that legitimately appears in two
    # categories collapses to one row and a REAL observed price is silently lost: Claro's Controle page
    # yields `claro:plano-flex-20gb` ("Controle 20GB", R$44,90) whose slug also names the Flex-page row,
    # so only 3 of the 4 Claro Controle plans were being stored (#35 audit).
    df = df.drop_duplicates(subset=["snapshot_date", "carrier", "state", "category", "_key"],
                            keep="last")
    df = df.sort_values(["snapshot_date", "carrier", "category", "plan_name"]).reset_index(drop=True)
    # Enforce the canonical schema/order (CONTEXT §6). This matters when merging a history written by
    # OLDER code that lacked a column (e.g. validity_days, #18): pd.concat would otherwise append the
    # new column at the END, so the matrix's column-letter formulas (which assume COLUMNS order) would
    # point at the wrong column. reindex also drops the temporary `_key`. (#18)
    return df.reindex(columns=COLUMNS)


def _rotation_pairs(cur, old, new_keys, gone_keys) -> dict:
    """Id-ROTATION fallback (#29): carriers republish plans with fresh native ids (TIM rotated every
    Drupal nid on 2026-07-06 while cutting Controle prices −15%), which turns real price changes into
    invisible "removed"+"new" pairs. Pair the id-unmatched rows by (carrier, state, category, plan_name)
    — names survive rotations — but ONLY when the name maps 1:1 on both sides (ambiguity → keep the
    honest new/removed rows). Returns {new_key: old_key}.

    ⚠️ TIGHTENED in #37, in lockstep with `alerts.compute_price_moves`. #29 only required the name to
    be unique among the UNMATCHED rows, so a name that ALSO exists on an id-matched row could still be
    claimed by the fallback (the #35 audit found 176 non-unique name groups and a false "Vivo Lite
    R$35 → R$75, +114%"). The name must now be unique in the FULL snapshot on BOTH sides. This helper
    and the alert matcher MUST stay identical: CONTEXT §3 promises the `changes` sheet and the e-mail
    agree, and a rule that lived in only one of them would make them contradict each other."""
    def name_key(r):
        return (str(r["carrier"]), str(r["state"]), str(r["category"]), str(r["plan_name"]))

    def by_name(frame, keys):
        m: dict = {}
        for k in keys:
            m.setdefault(name_key(frame.loc[k]), []).append(k)
        return m

    def counts(frame):                       # multiplicity across the WHOLE snapshot, not just leftovers
        c: dict = {}
        for _, r in frame.iterrows():
            k = name_key(r)
            c[k] = c.get(k, 0) + 1
        return c

    news, olds = by_name(cur, new_keys), by_name(old, gone_keys)
    cur_names, old_names = counts(cur), counts(old)
    return {ks[0]: olds[nk][0] for nk, ks in news.items()
            if len(ks) == 1 and len(olds.get(nk, [])) == 1
            and cur_names.get(nk, 0) == 1 and old_names.get(nk, 0) == 1}


def _merge_convergent(existing: pd.DataFrame, fresh: pd.DataFrame) -> pd.DataFrame:
    """Append-only merge for `convergent_history` (#31) — the same contract as `_merge_history`:
    a re-run of a snapshot_date REPLACES that date's rows (idempotent per date), other dates
    accumulate untouched. Identity = (snapshot_date, carrier, state, offer_id or offer_name).
    Reindexed to CONVERGENT_COLUMNS so an older sheet that lacks a column can't shift the order."""
    if not fresh.empty and "snapshot_date" in existing.columns and not existing.empty:
        existing = existing[~existing["snapshot_date"].isin(set(fresh["snapshot_date"].dropna().unique()))]
    df = pd.concat([existing, fresh], ignore_index=True)
    if df.empty:
        return df.reindex(columns=CONVERGENT_COLUMNS)
    oid = df["offer_id"].astype("string") if "offer_id" in df.columns else pd.Series(pd.NA, index=df.index, dtype="string")
    df["_key"] = oid.where(oid.notna() & (oid.str.strip() != ""), df["offer_name"].astype("string"))
    df = df.drop_duplicates(subset=["snapshot_date", "carrier", "state", "_key"], keep="last")
    df = df.sort_values(["snapshot_date", "carrier", "offer_name"]).reset_index(drop=True)
    df = _backfill_bundle_type(df)                    # pre-#33 rows gain the derived column
    return df.reindex(columns=CONVERGENT_COLUMNS)      # also drops the temporary `_key`


def _truthy(series: pd.Series) -> pd.Series:
    """Excel round-trips booleans as True/False, 1/0 or the strings "True"/"VERDADEIRO" — normalise."""
    if series.dtype == bool:
        return series
    return series.map(lambda v: str(v).strip().lower() in ("true", "1", "1.0", "verdadeiro", "sim", "yes"))


def _backfill_bundle_type(df: pd.DataFrame) -> pd.DataFrame:
    """Keep `bundle_type` correct and CURRENT for every row (#33, migration in #34).

    `bundle_type` is a pure function of the service flags, so where those flags are present we simply
    RECOMPUTE it — which fixes both cases in one rule:
      * rows written before the column existed (missing/NaN), and
      * rows stored under a superseded label — the "Fibre …" → "Fiber …" rename of #34. Preserving a
        stored value here (as #33 did) would leave historical rows spelled "Fibre + Mobile", they would
        stop matching the matrix's criteria, and every past date would silently render BLANK.
    When the flags are absent (an unexpectedly narrow sheet) we fall back to renaming the stored label
    rather than dropping it. Same self-healing contract as the mobile column additions in #18."""
    if df.empty:
        return df
    flags = ("has_mobile", "has_broadband", "has_tv", "has_landline")
    if not all(f in df.columns for f in flags):
        if "bundle_type" in df.columns:                      # no flags to derive from → migrate labels
            df["bundle_type"] = [migrate_bundle_type_label(v) for v in df["bundle_type"]]
        return df
    df["bundle_type"] = [
        bundle_type_of(m, b, t, l)
        for m, b, t, l in zip(*(_truthy(df[f]) for f in flags))
    ]
    return df


def _read_convergent(path: Path) -> pd.DataFrame:
    """Existing `convergent_history` rows, or an empty frame. NOTE (#31): `write_workbook` rebuilds
    the whole file (ExcelWriter mode 'w'), so a sheet that is not re-emitted is LOST — the convergent
    history is therefore ALWAYS read and re-written, even on a run that collected no convergent data
    (a failed/skipped convergent scrape must never wipe the convergent history)."""
    if not path.exists():
        return pd.DataFrame(columns=CONVERGENT_COLUMNS)
    try:
        return pd.read_excel(path, sheet_name="convergent_history", dtype={"snapshot_date": str})
    except Exception:
        return pd.DataFrame(columns=CONVERGENT_COLUMNS)    # sheet absent (pre-#31 workbook) or unreadable


def _compute_changes(history: pd.DataFrame) -> pd.DataFrame:
    cols = ["change_type", "carrier", "category", "state", "plan_id", "plan_name",
            "old_price_brl", "new_price_brl", "delta_brl"]
    dates = sorted(history["snapshot_date"].dropna().unique())
    if len(dates) < 2:
        return pd.DataFrame(columns=cols)
    latest, prev = dates[-1], dates[-2]
    h = _with_key(history)
    # `category` belongs in the match key for the same reason it belongs in the dedupe key (#37/C):
    # a native id is only unique WITHIN a category page. Claro's controle page reuses the flex slug,
    # so `claro:plano-flex-20gb` is a real R$44,90 Controle plan AND a real R$59,90 Flex plan in the
    # same snapshot — without `category` they collapse into one row and the R$15 gap reads as a move.
    h["_ck"] = (h["carrier"].astype(str) + "|" + h["state"].astype(str) + "|"
                + h["category"].astype(str) + "|" + h["_key"].astype(str))
    cur = h[h.snapshot_date == latest].drop_duplicates("_ck").set_index("_ck")
    old = h[h.snapshot_date == prev].drop_duplicates("_ck").set_index("_ck")

    def base(r):  # carrier, category, state, plan_id, plan_name — matched by (carrier, state, plan_id)
        return [r["carrier"], r["category"], r["state"], r["plan_id"], r["plan_name"]]

    new_keys = cur.index.difference(old.index)
    gone_keys = old.index.difference(cur.index)
    # id-rotation fallback (#29): unambiguous same-name pairs are the SAME plan re-keyed — compare
    # prices (a real move → price_change with the NEW id; same price → a pure re-key, no row at all).
    rotated = _rotation_pairs(cur, old, new_keys, gone_keys)
    consumed_old = set(rotated.values())

    out = []
    for k in new_keys:
        r = cur.loc[k]
        if k in rotated:
            o = old.loc[rotated[k]]
            np_, op = r["price_brl"], o["price_brl"]
            if pd.notna(np_) and pd.notna(op) and float(np_) != float(op):
                out.append(["price_change", *base(r), op, np_, float(np_) - float(op)])
            continue
        out.append(["new", *base(r), None, r["price_brl"], None])
    for k in gone_keys:
        if k in consumed_old:
            continue
        r = old.loc[k]
        out.append(["removed", *base(r), r["price_brl"], None, None])
    for k in cur.index.intersection(old.index):
        r, o = cur.loc[k], old.loc[k]
        np_, op = r["price_brl"], o["price_brl"]
        if pd.notna(np_) and pd.notna(op) and float(np_) != float(op):
            out.append(["price_change", *base(r), op, np_, float(np_) - float(op)])
    return pd.DataFrame(out, columns=cols)


def _format_sheet(ws, currency_cols=(), number_cols=(), tab_color=None, autofilter=False):
    _gridless(ws, tab_color)
    ws.freeze_panes = "A2"
    last_col = ws.max_column
    last_row = ws.max_row
    if last_col == 0:
        return
    if autofilter:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{max(last_row, 1)}"
    _house_header(ws, 1, last_col)
    for col in range(1, last_col + 1):
        letter = get_column_letter(col)
        width = max(10, min(40, len(str(ws.cell(row=1, column=col).value)) + 2))
        ws.column_dimensions[letter].width = width
        for row in range(2, last_row + 1):
            c = ws.cell(row=row, column=col)
            c.font = Font(name=FONT)
            if col in currency_cols:
                c.number_format = CURRENCY_FMT
                c.alignment = RIGHT
            elif col in number_cols:
                c.number_format = "#,##0.##"
                c.alignment = RIGHT
    if last_row >= 2:
        _band(ws, 2, last_row, last_col)


COMPARE_CARRIERS = ["vivo", "claro", "tim"]
# (group title, category values it spans, optional caveat note) — CONTEXT §7 methodology
COMPARE_GROUPS = [
    ("Pure Postpaid", ["postpaid"], None),
    ("Control / Hybrid", ["control"], None),
    ("Prepaid", ["prepaid"],
     "Note: prepaid is not a clean unit — Claro Prezão is a daily fee (R$1/dia) vs Vivo/TIM recharge amounts."),
    ("Digital", ["lite", "flex", "fit"],
     "Digital = Vivo Lite + Claro Flex + TIM Fit (#28); entry-level = no-commitment versions."),
]


def _ranked(latest: pd.DataFrame, cats, carrier: str):
    df = latest[(latest["carrier"] == carrier) & (latest["category"].isin(cats))]
    if "fit" in cats and "loyalty_months" in df.columns:
        # Digital ranking compares NO-COMMITMENT entry prices (the in-sheet note says so): a fit version
        # whose OWN headline price requires loyalty (the Anual, #28) would rank apples-to-oranges against
        # Vivo/Claro no-commitment prices — drop it HERE only (it stays in history + the TIM sheet).
        # Lite toggle cards already headline the no-commitment price (loyalty price is the promo), so
        # only fit needs this. (#28, adversarial-review fix)
        loy = pd.to_numeric(df["loyalty_months"], errors="coerce")
        df = df[~((df["category"] == "fit") & (loy > 0))]
    df = df.dropna(subset=["price_brl"]).sort_values("price_brl", kind="stable")
    return list(df[["price_brl", "plan_name", "data_gb", "price_promo_brl"]]
                .itertuples(index=False, name=None))


def build_comparison_data(latest: pd.DataFrame) -> list[dict]:
    """Pure: latest snapshot → per-group, per-carrier price-ranked plan lists (CONTEXT §7).
    Each carrier's plans in a category are sorted ascending by price_brl, then aligned across
    carriers by rank (cheapest-vs-cheapest, 2nd-vs-2nd, …). Offline-testable."""
    out = []
    for title, cats, note in COMPARE_GROUPS:
        per = {c: _ranked(latest, cats, c) for c in COMPARE_CARRIERS}
        out.append({
            "title": title,
            "note": note,
            "per_carrier": per,
            "max_rank": max((len(v) for v in per.values()), default=0),
        })
    return out


def _write_ranking(xl, latest: pd.DataFrame):
    """Render the `Ranking` sheet: the validated cross-section — four groups, each rank-aligned
    across Vivo/Claro/TIM by ascending price (today's snapshot)."""
    ws = xl.book.create_sheet("Ranking")
    _gridless(ws, TAB_COLORS["Ranking"])
    ws.freeze_panes = "A1"
    ws.cell(row=1, column=1,
            value="Ranking (today) — within category, aligned by price rank") \
        .font = Font(name=FONT, bold=True, size=14)
    headers = ["Rank", "Vivo R$", "Claro R$", "TIM R$", "Vivo plan", "Claro plan", "TIM plan"]
    r = 3
    for grp in build_comparison_data(latest):
        ws.cell(row=r, column=1, value=grp["title"]).font = Font(name=FONT, bold=True, size=12)
        r += 1
        if grp["note"]:
            cell = ws.cell(row=r, column=1, value=grp["note"])
            cell.font = Font(name=FONT, italic=True, size=9, color="808080")
            r += 1
        for j, h in enumerate(headers, start=1):
            ws.cell(row=r, column=j, value=h)
        _house_header(ws, r, len(headers))
        r += 1
        per = grp["per_carrier"]
        rank_top = r
        for rank in range(grp["max_rank"]):
            ws.cell(row=r, column=1, value=rank + 1).font = Font(name=FONT)
            for ci, carrier in enumerate(COMPARE_CARRIERS):
                rows = per[carrier]
                if rank < len(rows):
                    price, name, gb, promo = rows[rank]
                    pcell = ws.cell(row=r, column=2 + ci, value=float(price))
                    pcell.number_format = CURRENCY_FMT
                    pcell.font = Font(name=FONT)
                    pcell.alignment = RIGHT
                    label = f"{name} ({int(gb)}GB)" if pd.notna(gb) else str(name)
                    if pd.notna(promo):
                        label += f" · promo R${float(promo):g}"
                    ws.cell(row=r, column=5 + ci, value=label).font = Font(name=FONT)
            r += 1
        rank_bottom = r - 1
        if rank_bottom >= rank_top:
            _band(ws, rank_top, rank_bottom, len(headers))
            rng = f"B{rank_top}:D{rank_bottom}"        # the three R$ columns over this group's rank rows
            # 1) cheapest-of-the-row highlight (higher priority + stopIfTrue so it wins over the scale)
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND(B{rank_top}<>"",B{rank_top}=MIN($B{rank_top}:$D{rank_top}))'],
                fill=CHEAP_FILL, font=CHEAP_FONT, stopIfTrue=True))
            # 2) colour-grade the R$ columns green (cheap) → yellow → red (expensive)
            ws.conditional_formatting.add(rng, _price_scale())
        r += 1  # blank separator between groups
    for col, w in {1: 6, 2: 11, 3: 11, 4: 11, 5: 32, 6: 32, 7: 32}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


OPERATOR_DISPLAY = {"vivo": "Vivo", "claro": "Claro", "tim": "TIM"}
# category blocks within each operator sheet, in display order
OPERATOR_CATEGORY_ORDER = [
    ("Postpaid", ["postpaid"]),
    ("Control", ["control"]),
    ("Digital", ["lite", "flex", "fit"]),
    ("Prepaid", ["prepaid"]),
]
OPERATOR_COLUMNS = ["Plan", "Price R$", "Promo R$", "Data", "Voice",
                    "Unlimited apps", "Streaming", "Notes"]


def _s(v) -> str | None:
    """A non-empty string, else None (so blank cells stay blank)."""
    return v if isinstance(v, str) and v.strip() else None


def _operator_plan_row(p) -> list:
    """One readable catalog row from a latest-snapshot record (itertuples)."""
    data = f"{p.data_gb:g} GB" if pd.notna(p.data_gb) else ""
    dnote = _s(p.data_note)
    if dnote:
        data = f"{data} ({dnote})" if data else dnote
    vd = getattr(p, "validity_days", None)              # prepaid validity, e.g. "30d" (#18)
    if pd.notna(vd):
        tag = f"{int(vd)}d"
        data = f"{data} · {tag}" if data else tag
    notes = " · ".join(x for x in (_s(p.extra_benefits), _s(p.price_note)) if x) or None
    return [
        p.plan_name,
        float(p.price_brl) if pd.notna(p.price_brl) else None,
        float(p.price_promo_brl) if pd.notna(p.price_promo_brl) else None,
        data or None,
        _s(p.voice),
        _s(p.unlimited_apps),
        _s(p.streaming),
        notes,
    ]


def build_operator_sheets(latest: pd.DataFrame) -> list[dict]:
    """Pure: latest snapshot → one structure per carrier present (category blocks, price-sorted).
    Carriers not in the snapshot get no sheet. Offline-testable."""
    present = set(latest["carrier"].dropna())
    ordered = [c for c in COMPARE_CARRIERS if c in present]
    ordered += [c for c in latest["carrier"].dropna().unique() if c not in ordered]  # future carriers
    out = []
    for carrier in ordered:
        blocks = []
        for label, cats in OPERATOR_CATEGORY_ORDER:
            df = latest[(latest["carrier"] == carrier) & (latest["category"].isin(cats))]
            df = df.dropna(subset=["price_brl"]).sort_values("price_brl", kind="stable")
            rows = [_operator_plan_row(p) for p in df.itertuples(index=False)]
            if rows:
                blocks.append({"label": label, "rows": rows})
        if blocks:
            out.append({"carrier": carrier,
                        "display": OPERATOR_DISPLAY.get(carrier, str(carrier).title()),
                        "blocks": blocks})
    return out


def _write_operator_sheets(xl, latest: pd.DataFrame):
    """One sheet per carrier: plans grouped by category (accented sub-header), price-sorted."""
    ncol = len(OPERATOR_COLUMNS)
    for op in build_operator_sheets(latest):
        display = op["display"]
        ws = xl.book.create_sheet(display)
        _gridless(ws, TAB_COLORS.get(display))
        ws.freeze_panes = "A2"
        for j, h in enumerate(OPERATOR_COLUMNS, start=1):
            ws.cell(row=1, column=j, value=h)
        _house_header(ws, 1, ncol)
        r = 2
        for block in op["blocks"]:
            for col in range(1, ncol + 1):                # category sub-header band
                ws.cell(row=r, column=col).fill = SUBHEADER_FILL
            ws.cell(row=r, column=1, value=block["label"]).font = SUBHEADER_FONT
            r += 1
            block_top = r
            for i, row in enumerate(block["rows"]):
                for j, val in enumerate(row, start=1):
                    cell = ws.cell(row=r, column=j, value=val)
                    cell.font = Font(name=FONT)
                    if j in (2, 3):                        # Price R$, Promo R$
                        cell.number_format = CURRENCY_FMT
                        cell.alignment = RIGHT
                if i % 2 == 1:                             # banding within the block
                    for col in range(1, ncol + 1):
                        ws.cell(row=r, column=col).fill = BAND_FILL
                if row[2] is not None:                     # promo present → accent the Promo cell
                    ws.cell(row=r, column=3).fill = PROMO_FILL
                r += 1
            _color_scale(ws, 2, block_top, r - 1)          # grade Price R$ within this category block
        for col, w in {1: 32, 2: 11, 3: 11, 4: 22, 5: 20, 6: 18, 7: 18, 8: 40}.items():
            ws.column_dimensions[get_column_letter(col)].width = w


# ---- price-evolution matrix (the `comparison` sheet) — CONTEXT §7 ----------------------------
# (group title, category, kind). "single" = one category; "digital" = min over lite OR flex.
EVOLUTION_GROUPS = [
    ("Control (R$/mo)", "control", "single"),
    ("Post (R$/mo)", "postpaid", "single"),
    ("Pre (R$/mo, 30-day)", "prepaid", "prepaid30"),     # cheapest 30-day prepaid plan (#18)
    ("Digital (R$/mo)", None, "digital"),
]
# The Pre column compares the cheapest PREPAID plan whose validity_days >= this (a real 30-day plan,
# not the R$1/day or a 15-day tier). Mirrors config/sources.yaml `prepaid.min_validity_days`; the
# matrix is built without Settings, so the operative value is mirrored here (change both). #18
PREPAID_MIN_VALIDITY_DAYS = 28
EVOLUTION_CARRIERS = ["tim", "vivo", "claro"]
EVOLUTION_DISP = {"tim": "TIM", "vivo": "Vivo", "claro": "Claro"}
CARRIER_LINE = {"tim": "0033A0", "vivo": "660099", "claro": "DA291C"}  # series colors = tab palette
# Fixed Y-axis ranges per category (JPMorgan Figs 7/8/9 style, #25): a STABLE view with headroom so the
# axis doesn't jump around as daily data accumulates. Keyed by the short group title. Pre reflects the
# 30-day MONTHLY value we show (~R$20–40), NOT JPMorgan's R$/day (the Bridge's no-avg-daily call).
EVOLUTION_YRANGE = {"Control": (40, 70), "Post": (90, 170), "Pre": (20, 40), "Digital": (25, 65)}
# Table-only `comparison` layout (#17): group titles row 1, TIM/Vivo/Claro sub-header row 2, day rows 3+.
EVO_HEAD1, EVO_HEAD2, EVO_FIRST = 1, 2, 3


def _clean_gridlines(chart) -> None:
    """JPMorgan look (#25): NO vertical gridlines; exactly ONE faint horizontal gridline on the value
    axis (~0.5pt light grey). Works for both the line charts (value axis = y) and the grouped bar
    (type='col', value axis = y)."""
    chart.x_axis.majorGridlines = None                        # drop vertical gridlines
    gl = ChartLines(spPr=GraphicalProperties())
    gl.spPr.line = LineProperties(solidFill="ECECEC", w=6350)  # faint light-grey, ~0.5pt
    chart.y_axis.majorGridlines = gl


def _line_chart(chart_ws, data_ws, title, c0, head_row, first, last, anchor, yrange=None):
    """One per-category line chart hosted on `chart_ws` (the `Charts` sheet) but reading the matrix on
    `data_ws` (the `comparison` sheet) via CROSS-SHEET refs: X = the Date column, 3 series = the
    TIM/Vivo/Claro matrix columns. References point at the live matrix, so the charts grow as days
    accumulate; data_ws != chart_ws, so the table sheet stays chart-free (#17/#19). JPMorgan Figs 7/8/9
    styling (#25): bottom legend, carrier palette, markers, faint single horizontal gridline, and a FIXED
    Y range (`yrange`) so the view is stable as history grows."""
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Date"
    chart.height, chart.width = 7.2, 11.5
    chart.legend.position = "b"                  # legend at the bottom (clean report look, #20)
    _clean_gridlines(chart)                      # no vertical; one faint horizontal (#25)
    if yrange:                                   # fixed Y range → stable JPMorgan-style axis (#25)
        chart.y_axis.scaling.min, chart.y_axis.scaling.max = yrange
    data = Reference(data_ws, min_col=c0, max_col=c0 + 2, min_row=head_row, max_row=last)  # incl. names
    cats = Reference(data_ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for s, carrier in zip(chart.series, EVOLUTION_CARRIERS):
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(solidFill=CARRIER_LINE[carrier], w=20000)
        s.marker = Marker(symbol="circle", size=6)
        s.smooth = False
    chart_ws.add_chart(chart, anchor)
    return chart


def evolution_dates(history: pd.DataFrame) -> list[str]:
    """Distinct snapshot_dates in history, chronological (oldest first). Offline-testable."""
    if "snapshot_date" not in history.columns:
        return []
    return sorted({str(d) for d in history["snapshot_date"].dropna()})


def evolution_months(history: pd.DataFrame) -> list[date]:
    """Distinct MONTHS present in history, as first-of-month dates, chronological (oldest first).
    Built from snapshot_date (ISO text 'YYYY-MM-DD'), deduped to (year, month). The matrix has one
    row per month (the Figure-5 roll-up); daily detail stays in the `history` sheet. Offline-testable."""
    if "snapshot_date" not in history.columns:
        return []
    months = set()
    for d in history["snapshot_date"].dropna():
        s = str(d)
        try:
            months.add((int(s[0:4]), int(s[5:7])))
        except (ValueError, IndexError):
            continue
    return [date(y, m, 1) for (y, m) in sorted(months)]


def _hist_cols():
    return (get_column_letter(COLUMNS.index("price_brl") + 1),
            get_column_letter(COLUMNS.index("carrier") + 1),
            get_column_letter(COLUMNS.index("category") + 1),
            get_column_letter(COLUMNS.index("snapshot_date") + 1))


def _minifs_day(pc, cc, dc, sc, carrier, category, row, extra_crit=""):
    """Cheapest price `carrier` offered in `category` ON the EXACT date in $A{row}, computed live over
    the `history` sheet. The matrix is DAILY (#19): one row per snapshot_date. `extra_crit` appends one
    more MINIFS criteria pair (the Pre column adds validity_days >= 28, #18).

    Why a built text date and not the date cell / a >= bound: history stores snapshot_date as TEXT, and
    Excel coerces a date/serial comparison criterion to a number that never matches a text column —
    verified blank in real Excel (#16). So we match history's text date with an EXACT text date built
    from the row's DATE cell: YEAR()&"-"&TEXT(MONTH(),"00")&"-"&TEXT(DAY(),"00") → "2026-06-22" (no
    wildcard). The numeric "00" format keeps it locale-independent (NOT "yyyy"/"mm"/"dd" date codes,
    which pt-BR Excel localizes). (CONTEXT §7)"""
    daystr = (f'YEAR($A{row})&"-"&TEXT(MONTH($A{row}),"00")&"-"&TEXT(DAY($A{row}),"00")')
    return (f'_xlfn.MINIFS(history!${pc}:${pc},'
            f'history!${cc}:${cc},"{carrier}",'
            f'history!${dc}:${dc},"{category}",'
            f'history!${sc}:${sc},{daystr}'
            f'{extra_crit})')


def _matrix_value(history: pd.DataFrame, carrier: str, category, kind: str, day_str: str):
    """The Python value that the matrix cell's MINIFS computes — cheapest R$ for carrier/category on the
    exact date — used to BAKE a cached value alongside the formula (#21). Same logic as `_minifs_day`:
    single = min over the category; prepaid30 = min over prepaid with validity_days>=28; digital = min
    over lite+flex with loyalty_months>0 (#26). Postpaid = plain cheapest, credit-card-only INCLUDED
    (#27 reverted the #24 exclusion; the payment_method tag stays in history as context). Returns a
    rounded float, or None for no match (→ the cell stays blank, matching the formula's
    IF(=0,"") / IFERROR)."""
    h = history[(history["carrier"] == carrier) & (history["snapshot_date"].astype(str) == day_str)]
    if kind == "prepaid30":
        v = pd.to_numeric(h.get("validity_days"), errors="coerce")
        sub = h[(h["category"] == "prepaid") & (v >= PREPAID_MIN_VALIDITY_DAYS)]["price_brl"]
    elif kind == "digital":
        # Vivo Lite entry = the cheapest plan that OFFERS a "Sem fidelidade" choice (a monthly/annual
        # toggle card → parser sets loyalty_months). Excludes monthly-ONLY tiers like the 20GB "Plano
        # mensal" R$35 (no loyalty option) so the entry-level is the cheapest Sem-fidelidade plan (R$45),
        # per the Bridge (#26). Flex (Claro) has no loyalty toggle → unrestricted. Pre-#23 lite rows have
        # no loyalty_months → excluded → those dates self-heal to blank (like #18/#23).
        litep = h[h["category"] == "lite"]
        if "loyalty_months" in litep.columns:
            litep = litep[pd.to_numeric(litep["loyalty_months"], errors="coerce") > 0]
        # TIM Fit (#28) is the OPPOSITE shape: the two versions are SEPARATE plans, and the version whose
        # headline requires NO commitment is the Mensal (loyalty_months blank); the Anual's own price
        # (R$30) requires 12-mo permanence → excluded from the entry pick (kept in history).
        fitp = h[h["category"] == "fit"]
        if "loyalty_months" in fitp.columns:
            fitp = fitp[pd.to_numeric(fitp["loyalty_months"], errors="coerce").isna()]
        sub = pd.concat([litep["price_brl"], h[h["category"] == "flex"]["price_brl"],
                         fitp["price_brl"]])
    else:  # single — cheapest in the category, credit-card-only INCLUDED (#27 reverted the #24 exclusion)
        sub = h[h["category"] == category]["price_brl"]
    sub = pd.to_numeric(sub, errors="coerce").dropna()
    return round(float(sub.min()), 2) if len(sub) else None


def _write_comparison(xl, history: pd.DataFrame):
    """`comparison` = the price-evolution MATRIX, TABLE ONLY (#17), DAILY rows (#19): Date (one row per
    snapshot_date, earliest first) × category-group × carrier (cols), starting at the TOP of the sheet
    (titles row 1, TIM/Vivo/Claro row 2, days row 3+). Every value cell is a LIVE exact-date MINIFS over
    the `history` sheet, so the matrix grows one row per day and auto-populates as the daily job adds
    snapshots (self-connected workbook). 0 (no match) → "" so the heatmap ignores it. The 4 charts live
    on the separate `Charts` sheet (#17), so this sheet is a clean, scrollable table — only the 2 header
    rows + the Date column are frozen. Returns (worksheet, last_data_row, {coord: value}) — the value map
    is baked as cached values after save so the file displays in any viewer while keeping formulas (#21)."""
    ws = xl.book.create_sheet("comparison")
    _gridless(ws, TAB_COLORS["comparison"])
    pc, cc, dc, sc = _hist_cols()
    vc = get_column_letter(COLUMNS.index("validity_days") + 1)   # history validity_days column (#18)
    lc = get_column_letter(COLUMNS.index("loyalty_months") + 1)   # history loyalty_months column (#26)
    ncol = 1 + 3 * len(EVOLUTION_GROUPS)

    HEAD1, HEAD2, FIRST = EVO_HEAD1, EVO_HEAD2, EVO_FIRST   # table at the top: 1 / 2 / 3+
    ws.cell(row=HEAD1, column=1, value="Date")
    ws.merge_cells(start_row=HEAD1, start_column=1, end_row=HEAD2, end_column=1)
    for g, (title, _cat, _kind) in enumerate(EVOLUTION_GROUPS):
        c0 = 2 + g * 3
        ws.merge_cells(start_row=HEAD1, start_column=c0, end_row=HEAD1, end_column=c0 + 2)
        ws.cell(row=HEAD1, column=c0, value=title)
        for k, carrier in enumerate(EVOLUTION_CARRIERS):
            ws.cell(row=HEAD2, column=c0 + k, value=EVOLUTION_DISP[carrier])
    _house_header(ws, HEAD1, ncol)
    _house_header(ws, HEAD2, ncol)
    ws.cell(row=HEAD1, column=1).alignment = Alignment(vertical="center", horizontal="left")
    for c in range(2, ncol + 1):
        ws.cell(row=HEAD1, column=c).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=HEAD2, column=c).alignment = Alignment(vertical="center", horizontal="center")

    r = FIRST
    comp_values: dict[str, float] = {}                     # {coord: value} to bake as cached values (#21)
    for d_str in evolution_dates(history):                  # one row per snapshot_date, earliest first (#19)
        try:
            day = date.fromisoformat(str(d_str)[:10])
        except (ValueError, TypeError):
            continue
        dcell = ws.cell(row=r, column=1, value=day)
        dcell.number_format = "dd-mmm"                      # 2026-06-22 → "22-Jun"
        dcell.font = Font(name=FONT)
        for g, (title, category, kind) in enumerate(EVOLUTION_GROUPS):
            c0 = 2 + g * 3
            for k, carrier in enumerate(EVOLUTION_CARRIERS):
                if kind == "single":
                    # Post = the CHEAPEST postpaid plan, credit-card-only INCLUDED (#27 — Bridge reverted
                    # the #24 exclusion): TIM's R$119,99 "Express" (no cartão) IS the tracked entry-level;
                    # a change on it is what we want to detect (the changes sheet + the #15 alert watch its
                    # plan_id). The payment_method TAG stays in history (#24) so credit-card vs bill is
                    # still visible per plan — only the matrix pick stopped excluding it.
                    m = _minifs_day(pc, cc, dc, sc, carrier, category, r)
                    formula = f'=IFERROR(IF({m}=0,"",{m}),"")'
                elif kind == "prepaid30":   # cheapest 30-day prepaid plan (validity_days >= 28) — #18
                    crit = f',history!${vc}:${vc},">={PREPAID_MIN_VALIDITY_DAYS}"'
                    m = _minifs_day(pc, cc, dc, sc, carrier, "prepaid", r, extra_crit=crit)
                    formula = f'=IFERROR(IF({m}=0,"",{m}),"")'
                else:  # digital = cheapest of lite OR flex OR fit that day; "" if the carrier has none.
                    # reciprocal-max picks the smaller positive price: a day's MINIFS returns 0 on
                    # no-match (and no real plan is priced 0), so 1/0 → IFERROR → 0, and MAX keeps the
                    # larger reciprocal = the smaller price; all absent → MAX(0,…)=0 → 1/0 → "".
                    # LITE requires loyalty_months > 0 (#26): only plans that OFFER a "Sem fidelidade" choice
                    # (a monthly/annual toggle card) count, so the monthly-ONLY 20GB "Plano mensal" R$35 is
                    # excluded and the entry-level is the cheapest Sem-fidelidade plan (Vivo → R$45). Excel
                    # numeric ">0" excludes blanks, so pre-#23 lite rows (no loyalty) self-heal to blank.
                    lite = _minifs_day(pc, cc, dc, sc, carrier, "lite", r,
                                       extra_crit=f',history!${lc}:${lc},">0"')
                    flex = _minifs_day(pc, cc, dc, sc, carrier, "flex", r)   # Flex has no toggle → unrestricted
                    # FIT (TIM, #28) is the OPPOSITE of lite: its versions are SEPARATE plans, and only the
                    # no-commitment Mensal (loyalty_months BLANK — criterion "=") sets the entry; the Anual's
                    # own price requires 12-mo permanence → excluded (stays in history/the TIM sheet).
                    fit = _minifs_day(pc, cc, dc, sc, carrier, "fit", r,
                                      extra_crit=f',history!${lc}:${lc},"="')
                    formula = (f'=IFERROR(1/MAX(IFERROR(1/{lite},0),IFERROR(1/{flex},0),'
                               f'IFERROR(1/{fit},0)),"")')
                cell = ws.cell(row=r, column=c0 + k, value=formula)
                cell.number_format = CURRENCY_FMT
                cell.font = Font(name=FONT)
                cell.alignment = RIGHT
                v = _matrix_value(history, carrier, category, kind, d_str)   # bake cached value (#21)
                if v is not None:
                    comp_values[f"{get_column_letter(c0 + k)}{r}"] = v
        r += 1
    last = r - 1

    if last >= FIRST:    # per-category-BLOCK heatmap: one green→yellow scale over each group's 3 carrier
        for g in range(len(EVOLUTION_GROUPS)):   # columns × all date rows, so colour reflects price vs
            c0 = 2 + g * 3                        # ALL carriers in that category (JPMorgan style, #20)
            rng = f"{get_column_letter(c0)}{FIRST}:{get_column_letter(c0 + 2)}{last}"
            ws.conditional_formatting.add(rng, _price_scale_2color())

    note = ws.cell(row=last + 2, column=1,
                   value="Rows = day: each cell = cheapest R$ that carrier offered in the category on "
                         "that snapshot_date — a live MINIFS over `history` matching the EXACT date (a "
                         "locale-safe text date built from the row's date, since history stores dates as "
                         "TEXT — CONTEXT §7). Pre = cheapest 30-day prepaid plan (validity_days >= 28; "
                         "pre-#18 dates have no validity → blank, fills from 26-Jun on). Digital = min of "
                         "Vivo Lite / Claro Flex / TIM Fit — no-commitment versions only (lite: the Sem-"
                         "fidelidade price of toggle cards, #26; fit: the Mensal version, #28). Heatmap = "
                         "per-category-block green→yellow (each cell vs "
                         "all carriers in its category; subtle while prices are near-flat, differentiates "
                         "as they move). Grows one row per day; charts on the `Charts` sheet.")
    note.font = Font(name=FONT, italic=True, size=9, color="808080")
    # TIM Post methodology note (#27, reverting #24's exclusion per the Bridge): the Post column tracks the
    # CHEAPEST postpaid plan — for TIM that is the R$119,99 credit-card-only "Express" plan — so any change
    # on it is caught by the changes sheet + the daily alert. payment_method in history keeps the context.
    tim_note = ws.cell(row=last + 3, column=1,
                       value="Post — TIM: the entry-level tracked is the cheapest postpaid plan, "
                             "R$119,99 — note it is CREDIT-CARD-only (\"no cartão\"; TIM's Express line); "
                             "the cheapest bill-payment (\"na fatura\") plan is R$129,99. Both are in "
                             "`history` + the TIM sheet, with the billing type in `payment_method` — a "
                             "price change on the tracked plan shows in `changes` and the daily alert.")
    tim_note.font = Font(name=FONT, italic=True, size=9, color="808080")
    ws.freeze_panes = f"B{FIRST}"      # freeze only the 2 header rows + the Date column (scrolls freely)
    ws.protection.sheet = False        # explicit: the table is never locked (the #17 "locked view" fix)
    ws.column_dimensions["A"].width = 13
    for c in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10
    return ws, last, comp_values


def _write_charts(xl, comp_ws, first: int, last: int):
    """`Charts` sheet (#17): the 4 per-category line charts (Control / Post / Pre / Digital) in a 2×2
    layout + a current-price grouped bar chart below them (#20), each referencing the `comparison` matrix
    via cross-sheet refs (so they grow / update as days accumulate, #19). Clean report styling — bottom
    legend, no vertical gridlines, carrier palette. Lives apart from the table so the `comparison` sheet
    opens as a clean, scrollable matrix with no chart overlay."""
    ws = xl.book.create_sheet("Charts")
    _gridless(ws, TAB_COLORS["Charts"])
    ws.cell(row=1, column=1,
            value="Price evolution — cheapest R$ per carrier × category, by day (source: `comparison`)") \
        .font = Font(name=FONT, bold=True, size=14)
    if last >= first:                                  # nothing to chart until ≥1 day exists
        anchors = ["A3", "J3", "A20", "J20"]           # 2×2 block of line charts
        for g, (title, _c, _k) in enumerate(EVOLUTION_GROUPS):
            short = title.split(" (")[0]               # "Control" / "Post" / "Pre" / "Digital"
            _line_chart(ws, comp_ws, f"{short} – Monthly Prices",     # JPMorgan Figs 7/8/9 title (#25)
                        2 + g * 3, EVO_HEAD2, first, last, anchors[g],
                        yrange=EVOLUTION_YRANGE.get(short))
        _current_price_bar(ws, comp_ws, last)          # current-price grouped bar chart (#20)
    return ws


def _current_price_bar(ws, comp_ws, last):
    """Grouped BAR chart of the MOST RECENT day's entry-level price per carrier × category (JPMorgan
    Fig-11 style, #20): X = Control / Post / Pre / Digital, 3 bars each = TIM / Vivo / Claro. Values read
    the LAST matrix row via cross-sheet refs held in a small helper table (off to the right at P37) — so
    they are structure-driven (no hardcoded numbers) and rebuilt each run to the current last row, auto-
    updating as new days land. Pre bar = the matrix's 30-day monthly Pre value."""
    comp = comp_ws.title
    HCOL, HROW = 16, 37        # helper table top-left = P37, clear of the bar chart (anchored at A37)
    ws.cell(row=HROW, column=HCOL, value="Latest-day price (bar source)") \
        .font = Font(name=FONT, italic=True, size=9, color="808080")
    for k, carrier in enumerate(EVOLUTION_CARRIERS):                       # header: TIM | Vivo | Claro
        ws.cell(row=HROW, column=HCOL + 1 + k, value=EVOLUTION_DISP[carrier]) \
            .font = Font(name=FONT, bold=True)
    for g, (title, _c, _k) in enumerate(EVOLUTION_GROUPS):                 # one row per category
        rr = HROW + 1 + g
        ws.cell(row=rr, column=HCOL, value=title.split(" (")[0]).font = Font(name=FONT)
        for k in range(3):
            col = get_column_letter(2 + g * 3 + k)        # matrix column for this group × carrier
            ref = f"'{comp}'!{col}{last}"
            # a no-offer cell (e.g. TIM has no Digital line → blank matrix cell) → NA() so the bar is a
            # GAP, not a misleading labelled R$ 0.00 bar; a real price passes through unchanged. (#20)
            c = ws.cell(row=rr, column=HCOL + 1 + k, value=f'=IF({ref}="",NA(),{ref})')
            c.number_format = CURRENCY_FMT
            c.font = Font(name=FONT)
    cat_top, cat_bot = HROW + 1, HROW + len(EVOLUTION_GROUPS)
    ws.column_dimensions[get_column_letter(HCOL)].width = 12        # readable source table (no #####)
    for k in range(3):
        ws.column_dimensions[get_column_letter(HCOL + 1 + k)].width = 11
    chart = BarChart()
    chart.type = "col"                                 # vertical grouped columns
    chart.grouping = "clustered"
    chart.title = "Current entry-level price by carrier (R$)"
    chart.y_axis.title = "R$"
    chart.height, chart.width = 8.0, 16.0
    chart.legend.position = "b"
    _clean_gridlines(chart)                            # no vertical; one faint horizontal (#25)
    data = Reference(ws, min_col=HCOL + 1, max_col=HCOL + 3, min_row=HROW, max_row=cat_bot)  # incl names
    cats = Reference(ws, min_col=HCOL, min_row=cat_top, max_row=cat_bot)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for s, carrier in zip(chart.series, EVOLUTION_CARRIERS):   # bar fill = carrier palette
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.solidFill = CARRIER_LINE[carrier]
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True                    # data labels on
    ws.add_chart(chart, "A37")
    return chart


# ---- convergent offers (#31) — a SEPARATE domain: combos (mobile + fiber [+ TV/landline]) ------------
def _write_convergent_history(xl, conv: pd.DataFrame):
    """`convergent_history` = append-only combo offers, one row per offer per snapshot_date (CONTEXT
    §14). Deliberately a PLAIN, flat sheet next to the mobile sheets — no formulas, no cached-value
    bake, no charts — so nothing here can perturb the mobile matrix/bake/charts. Written on EVERY run
    (even when empty) because ExcelWriter rebuilds the file; an empty frame still emits the headers so
    the sheet exists and the schema is visible from day one."""
    df = conv if len(conv) else pd.DataFrame(columns=CONVERGENT_COLUMNS)
    # DATA FIRST, unguarded and unadorned: this is the one call that must not be skipped, because a
    # missing sheet means the accumulated convergent history is gone on the NEXT read (the file is
    # rebuilt each run). Cosmetics are applied separately and are allowed to fail. (review #31)
    df.to_excel(xl, sheet_name="convergent_history", index=False)
    ws = xl.sheets["convergent_history"]
    try:
        cur = (CONVERGENT_COLUMNS.index("price_brl") + 1, CONVERGENT_COLUMNS.index("price_promo_brl") + 1)
        num = (CONVERGENT_COLUMNS.index("broadband_speed_mbps") + 1,
               CONVERGENT_COLUMNS.index("mobile_gb") + 1,
               CONVERGENT_COLUMNS.index("mobile_lines") + 1)
        _format_sheet(ws, cur, num, tab_color=TAB_COLORS["convergent_history"], autofilter=True)
    except Exception as e:      # styling only — the rows are already in the sheet
        print(f"convergent: sheet formatting failed ({type(e).__name__}: {e}) - data written unstyled")
    return ws


def convergent_bundle_types(conv: pd.DataFrame) -> list[str]:
    """Every bundle type PRESENT in the collected history, canonical order first then any new shape.
    This is the "what did we collect" view — used for the matrix's footnote, so a reader can see which
    types exist but are not charted. The matrix itself shows `COMPARISON_BUNDLE_TYPES` (#34)."""
    if conv.empty or "bundle_type" not in conv.columns:
        return []
    seen = {str(b).strip() for b in conv["bundle_type"].dropna() if str(b).strip()}
    return ([t for t in CANONICAL_BUNDLE_TYPES if t in seen]
            + sorted(seen - set(CANONICAL_BUNDLE_TYPES)))


def _conv_cols():
    """(price, carrier, bundle_type, snapshot_date) column letters on the convergent_history sheet."""
    return (get_column_letter(CONVERGENT_COLUMNS.index("price_brl") + 1),
            get_column_letter(CONVERGENT_COLUMNS.index("carrier") + 1),
            get_column_letter(CONVERGENT_COLUMNS.index("bundle_type") + 1),
            get_column_letter(CONVERGENT_COLUMNS.index("snapshot_date") + 1))


def _conv_minifs_day(pc, cc, bc, sc, carrier: str, bundle: str, row: int) -> str:
    """Cheapest price `carrier` offered for `bundle` ON the EXACT date in $A{row}, live over
    `convergent_history`. Same locale-safe TEXT-date construction as the mobile matrix (§7): the sheet
    stores snapshot_date as TEXT (verified), so a date/serial criterion would never match — we rebuild
    the ISO string from the row's date cell with numeric "00" formats (NOT localized date codes)."""
    daystr = (f'YEAR($A{row})&"-"&TEXT(MONTH($A{row}),"00")&"-"&TEXT(DAY($A{row}),"00")')
    return (f'_xlfn.MINIFS(convergent_history!${pc}:${pc},'
            f'convergent_history!${cc}:${cc},"{carrier}",'
            f'convergent_history!${bc}:${bc},"{bundle}",'
            f'convergent_history!${sc}:${sc},{daystr})')


def _conv_matrix_value(conv: pd.DataFrame, carrier: str, bundle: str, day_str: str):
    """The Python value the cell's MINIFS computes — used to BAKE a cached value beside the formula
    (#21 discipline: this MUST mirror `_conv_minifs_day` exactly). None when the carrier had no offer
    of that bundle type that day → the cell stays BLANK (an honest gap, never a misleading 0)."""
    if conv.empty or "bundle_type" not in conv.columns:
        return None
    sub = conv[(conv["carrier"] == carrier)
               & (conv["bundle_type"].astype("string") == bundle)
               & (conv["snapshot_date"].astype(str) == day_str)]["price_brl"]
    sub = pd.to_numeric(sub, errors="coerce").dropna()
    return round(float(sub.min()), 2) if len(sub) else None


def _write_convergent_comparison(xl, conv: pd.DataFrame):
    """`convergent_comparison` (#33, scoped in #34) — the combo equivalent of the mobile `comparison`
    matrix (§7). Rows = one per snapshot_date in `convergent_history` (earliest first, auto-growing);
    columns = the bundle types in `COMPARISON_BUNDLE_TYPES`, each split TIM | Vivo | Claro. Every value
    cell is a LIVE exact-date MINIFS over `convergent_history`, so the sheet fills itself as the daily
    job adds snapshots, and the computed values are baked alongside (#21) so the committed file
    displays in any viewer.

    SCOPE (#34): the matrix shows **Fiber + Mobile only** — the like-for-like headline bundle every
    carrier sells. All other bundle types are still COLLECTED into `convergent_history` (collection is
    never filtered); they are simply not charted here, and the footnote names the ones on file.
    Restoring a group is a one-line change to `COMPARISON_BUNDLE_TYPES` — no literals live here.

    Why cheapest WITHIN a bundle type rather than cheapest overall: carriers sell very different
    things under one "combo" banner — a fiber+TV bundle is not a competitor to a fiber+mobile one, and
    a single "cheapest combo" column would compare unlike with unlike. Returns (worksheet, {coord:
    value}) — an empty history yields a header-only sheet, never an exception."""
    ws = xl.book.create_sheet("convergent_comparison")
    _gridless(ws, TAB_COLORS["convergent_comparison"])
    groups = list(COMPARISON_BUNDLE_TYPES)      # the ONE knob (#34) — collection stays unfiltered
    pc, cc, bc, sc = _conv_cols()
    ncol = 1 + 3 * len(groups)

    HEAD1, HEAD2, FIRST = 1, 2, 3
    ws.cell(row=HEAD1, column=1, value="Date")
    ws.merge_cells(start_row=HEAD1, start_column=1, end_row=HEAD2, end_column=1)
    for g, title in enumerate(groups):
        c0 = 2 + g * 3
        ws.merge_cells(start_row=HEAD1, start_column=c0, end_row=HEAD1, end_column=c0 + 2)
        ws.cell(row=HEAD1, column=c0, value=f"{title} (R$/mo)")
        for k, carrier in enumerate(EVOLUTION_CARRIERS):
            ws.cell(row=HEAD2, column=c0 + k, value=EVOLUTION_DISP[carrier])
    _house_header(ws, HEAD1, ncol)
    _house_header(ws, HEAD2, ncol)
    ws.cell(row=HEAD1, column=1).alignment = Alignment(vertical="center", horizontal="left")
    for c in range(2, ncol + 1):
        ws.cell(row=HEAD1, column=c).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=HEAD2, column=c).alignment = Alignment(vertical="center", horizontal="center")

    values: dict[str, float] = {}
    r = FIRST
    for d_str in evolution_dates(conv):                 # one row per snapshot_date, earliest first
        try:
            day = date.fromisoformat(str(d_str)[:10])
        except (ValueError, TypeError):
            continue
        dcell = ws.cell(row=r, column=1, value=day)
        dcell.number_format = "dd-mmm"
        dcell.font = Font(name=FONT)
        for g, bundle in enumerate(groups):
            c0 = 2 + g * 3
            for k, carrier in enumerate(EVOLUTION_CARRIERS):
                m = _conv_minifs_day(pc, cc, bc, sc, carrier, bundle, r)
                cell = ws.cell(row=r, column=c0 + k, value=f'=IFERROR(IF({m}=0,"",{m}),"")')
                cell.number_format = CURRENCY_FMT
                cell.font = Font(name=FONT)
                cell.alignment = RIGHT
                v = _conv_matrix_value(conv, carrier, bundle, d_str)
                if v is not None:
                    values[f"{get_column_letter(c0 + k)}{r}"] = v
        r += 1
    last = r - 1

    if last >= FIRST:      # per-bundle-type-BLOCK heatmap: one green→yellow scale over each group's
        for g in range(len(groups)):     # 3 carrier columns × all date rows (same as the mobile #20)
            c0 = 2 + g * 3
            rng = f"{get_column_letter(c0)}{FIRST}:{get_column_letter(c0 + 2)}{last}"
            ws.conditional_formatting.add(rng, _price_scale_2color())

    others = [t for t in convergent_bundle_types(conv) if t not in groups]
    shown = " / ".join(groups)
    note = ws.cell(row=max(last, FIRST) + 2, column=1,
                   value=f"Rows = day. Each cell = the CHEAPEST {shown} combo that carrier offered on "
                         "that snapshot_date — a live MINIFS over `convergent_history` matching the "
                         "EXACT date (locale-safe text date, as in `comparison`). Compared WITHIN a "
                         "bundle type so the columns are like-for-like (a fiber+TV bundle is not a "
                         "competitor to a fiber+mobile one). A BLANK cell means that carrier sold no "
                         f"such combo that day — an honest gap, not a zero."
                         + (f" Other bundle types ARE still collected in `convergent_history` "
                            f"({', '.join(others)}) — they are just not charted here."
                            if others else "")
                         + " Bundle type is derived from the offer's service flags; the sheet grows "
                           "one row per day.")
    note.font = Font(name=FONT, italic=True, size=9, color="808080")
    ws.freeze_panes = f"B{FIRST}"
    ws.column_dimensions["A"].width = 13
    for c in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    return ws, values


# ---- cached-value bake (#21) — so the committed workbook DISPLAYS in any viewer, keeping formulas ------
_SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _sheet_xml_targets(zf: zipfile.ZipFile) -> dict:
    """{sheet display name: 'xl/worksheets/sheetN.xml'} via workbook.xml + its rels (read-only)."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid = {r.get("Id"): r.get("Target") for r in rels}
    out = {}
    for sh in wb.find(f"{{{_SS_NS}}}sheets"):
        t = (rid.get(sh.get(f"{{{_REL_NS}}}id")) or "").lstrip("/")
        if t:
            out[sh.get("name")] = t if t.startswith("xl/") else "xl/" + t
    return out


def _inject_cached_values(xml: str, coords: dict):
    """Set ``<v>value</v>`` for each named formula cell, ROBUST to openpyxl's serialization of a formula
    cell (#21). openpyxl 3.1.x writes ``<f>…</f>`` followed by an EMPTY value placeholder — ``<v/>`` (lxml
    build) OR ``<v></v>`` (stdlib etree build); older builds write no ``<v>`` at all. We DROP any empty
    placeholder and write ``<v>value</v>`` after ``</f>`` (so there's exactly one ``<v>``, never a
    duplicate); a cell that already carries a REAL value is left unchanged (idempotent). A surgical
    per-cell string edit — no XML re-serialization, so namespaces, drawings (charts), conditional
    formatting and every other cell stay byte-for-byte intact. Returns (xml, n_injected)."""
    n = 0
    for coord, val in coords.items():
        pat = re.compile(rf'(<c r="{re.escape(coord)}"[^>]*>)(.*?)(</c>)', re.S)

        def repl(m, val=val):
            nonlocal n
            inner = m.group(2)
            if "<f" not in inner or "</f>" not in inner:
                return m.group(0)                     # not a formula cell
            if re.search(r"<v>[^<]+</v>", inner):
                return m.group(0)                     # already has a real cached value (idempotent)
            inner = re.sub(r"<v\s*/>|<v>\s*</v>", "", inner)      # drop empty <v/> / <v></v> placeholder
            sval = str(int(val)) if float(val).is_integer() else str(val)   # 150.0 → "150", 58.99 → "58.99"
            n += 1
            return m.group(1) + inner.replace("</f>", f"</f><v>{sval}</v>", 1) + m.group(3)

        xml = pat.sub(repl, xml, count=1)
    return xml, n


def _bake_cached_values(path, value_map: dict) -> int:
    """Bake cached values into the named formula cells so the workbook DISPLAYS its values in ANY viewer
    (GitHub preview, Google Sheets, mobile, Excel Protected View) while KEEPING the live formulas.
    openpyxl writes ``<f>`` without ``<v>``, so those viewers (which don't recalc) show blank — this
    injects ``<v>`` alongside ``<f>``. Only the named cells are touched (charts / CF / formatting are
    left intact). Blank cells (no offer) are not in the map, so they correctly stay blank. Returns the
    number of cells baked. (#21)"""
    path = Path(path)
    with zipfile.ZipFile(path) as zf:
        targets = _sheet_xml_targets(zf)
        blobs = {n: zf.read(n) for n in zf.namelist()}
    baked = 0
    for sheet, coords in value_map.items():
        tgt = targets.get(sheet)
        if not tgt or tgt not in blobs or not coords:
            continue
        new_xml, n = _inject_cached_values(blobs[tgt].decode("utf-8"), coords)
        blobs[tgt] = new_xml.encode("utf-8")
        baked += n                                    # count ACTUAL injections (0-bake regressions detectable)
    tmp = path.with_name(path.name + ".bake.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in blobs.items():
            zf.writestr(name, data)
    os.replace(tmp, path)
    return baked


def write_workbook(plans, path: str | Path, run_ts: datetime, convergent=None) -> dict:
    """Rebuild the workbook from `plans` (mobile) — and, additively, the `convergent_history` sheet
    (#31). `convergent` is an optional list of ConvergentOffer; None/empty means "collected nothing
    this run", which PRESERVES the existing convergent history rather than dropping it (the sheet is
    always read + re-emitted because ExcelWriter rebuilds the file)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fresh = _df(plans)

    existing = pd.DataFrame(columns=COLUMNS)
    if path.exists():
        try:
            existing = pd.read_excel(path, sheet_name="history", dtype={"snapshot_date": str})
        except Exception:
            existing = pd.DataFrame(columns=COLUMNS)

    # Convergent (#31) — read ALWAYS (see _read_convergent), merge only what this run collected.
    # GUARDED: convergent is an additive side-domain, so a problem building it must degrade to
    # "keep what we already had" and NEVER fail the mobile workbook write (main.py guards the scrape;
    # this guards the write path itself — collection of mobile data > convergent).
    try:
        conv_fresh = pd.DataFrame([o.as_row() for o in (convergent or [])], columns=CONVERGENT_COLUMNS)
        conv_history = _merge_convergent(_read_convergent(path), conv_fresh)
    except Exception as e:                                    # noqa: BLE001 - must not break the run
        print(f"convergent: merge failed ({type(e).__name__}: {e}) - keeping existing rows")
        try:
            conv_history = _read_convergent(path)
        except Exception:
            conv_history = pd.DataFrame(columns=CONVERGENT_COLUMNS)

    history = _merge_history(existing, fresh)
    latest_date = history["snapshot_date"].max()
    latest = history[history.snapshot_date == latest_date].reset_index(drop=True)
    changes = _compute_changes(history)

    # price columns -> currency; data_gb -> number. (1-based indices from COLUMNS)
    cur_cols = (COLUMNS.index("price_brl") + 1, COLUMNS.index("price_promo_brl") + 1)
    num_cols = (COLUMNS.index("data_gb") + 1,)
    conv_written = True                     # flipped by the last-resort guard below (#31)
    conv_values: dict[str, float] = {}       # convergent_comparison cached values to bake (#33)

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        history.to_excel(xl, sheet_name="history", index=False)
        latest.to_excel(xl, sheet_name="latest", index=False)
        changes.to_excel(xl, sheet_name="changes", index=False)
        _format_sheet(xl.sheets["history"], cur_cols, num_cols,
                      tab_color=TAB_COLORS["history"], autofilter=True)
        latest_ws = xl.sheets["latest"]
        _format_sheet(latest_ws, cur_cols, num_cols, tab_color=TAB_COLORS["latest"], autofilter=True)
        if len(latest) >= 1:
            last = len(latest) + 1
            _color_scale(latest_ws, COLUMNS.index("price_brl") + 1, 2, last)        # price green→red
            L = get_column_letter(COLUMNS.index("price_promo_brl") + 1)
            latest_ws.conditional_formatting.add(                                    # promo accent
                f"{L}2:{L}{last}", FormulaRule(formula=[f"NOT(ISBLANK({L}2))"], fill=PROMO_FILL))
        _format_sheet(xl.sheets["changes"], (7, 8, 9), tab_color=TAB_COLORS["changes"])
        summ_values = _write_summary(xl, run_ts, latest_date, len(latest),
                                     history["snapshot_date"].nunique(), latest)
        comp_ws, comp_last, comp_values = _write_comparison(xl, history)   # table-only daily matrix
        _write_charts(xl, comp_ws, EVO_FIRST, comp_last)      # 4 line charts + bar chart on `Charts` (#17/#20)
        _write_ranking(xl, latest)           # validated cross-section (rank-aligned, today)
        _write_operator_sheets(xl, latest)   # one tab per carrier (by-operator catalog view)
        try:                                          # combo offers — separate domain, additive (#31)
            _write_convergent_history(xl, conv_history)
            # …and its evolution matrix (#33). Guarded SEPARATELY so a matrix problem cannot cost us
            # the collected history sheet, and so `conv_values` stays empty on failure — otherwise the
            # fail-loud bake check below would abort the whole (mobile) write over a side-domain bug.
            try:
                _, conv_values = _write_convergent_comparison(xl, conv_history)
            except Exception as e:
                conv_values = {}
                print(f"convergent: comparison sheet failed ({type(e).__name__}: {e}) - "
                      f"history kept, matrix skipped")
        except Exception as e:                        # never let the side-domain break the mobile write
            # Last-resort guard. The sheet is now absent from THIS workbook, so the accumulated
            # convergent rows will not be re-readable next run — say so LOUDLY rather than reporting
            # a count that was never persisted (the mobile snapshot is unaffected and still commits).
            conv_written = False
            print(f"convergent: SHEET WRITE FAILED ({type(e).__name__}: {e}) — convergent_history is "
                  f"MISSING from this workbook and {len(conv_history)} row(s) were not persisted; "
                  f"the mobile data is unaffected")
        # Open on the clean `comparison` table, not on `history`: set it active + the only selected tab.
        comp_idx = xl.book.index(comp_ws)
        xl.book.active = comp_idx
        for i, s in enumerate(xl.book.worksheets):
            s.sheet_view.tabSelected = (i == comp_idx)

    # openpyxl writes formulas WITHOUT cached values, so non-recalc viewers (GitHub preview, Google Sheets,
    # mobile, Excel Protected View) show blank. Bake the computed values into the formula cells — keeping
    # the live formulas — so the committed workbook displays everywhere. (#21, CONTEXT §5/§7)
    value_map = {"comparison": comp_values, "summary": summ_values,
                 "convergent_comparison": conv_values}                                    # (#33)
    baked = _bake_cached_values(path, value_map)
    expected = sum(len(v) for v in value_map.values())
    # Fail LOUDLY on a 0/partial bake (e.g. a future openpyxl serialization change defeating the injector)
    # rather than silently committing a workbook that displays blank in non-recalc viewers. (#21)
    if expected and baked != expected:
        raise RuntimeError(
            f"cached-value bake wrote {baked}/{expected} cells — the workbook would display blank in "
            f"non-recalc viewers (openpyxl serialization changed?). Aborting to avoid shipping it.")

    return {
        "path": str(path),
        "snapshot_date": str(latest_date),
        "plans_in_latest": int(len(latest)),
        "snapshots_in_history": int(history["snapshot_date"].nunique()),
        "changes": int(len(changes)),
        # counts describe what was actually PERSISTED to the sheet (0 if the write failed) — #31
        "convergent_rows": int(len(conv_history)) if conv_written else 0,
        "convergent_snapshots": (int(conv_history["snapshot_date"].nunique())
                                 if conv_written and len(conv_history) else 0),
    }


def _write_summary(xl, run_ts, latest_date, n_latest, n_snapshots, latest):
    wb = xl.book
    ws = wb.create_sheet("summary")
    _gridless(ws, TAB_COLORS["summary"])
    ws["A1"] = "Mobile Price Tracker — Summary"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    meta = [
        ("Run timestamp", run_ts.isoformat(timespec="seconds")),
        ("Latest snapshot", str(latest_date)),
        ("Plans in latest", n_latest),
        ("Snapshots in history", n_snapshots),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(name=FONT, bold=True)
        ws.cell(row=i, column=2, value=v).font = Font(name=FONT)

    # Per-carrier KPI table driven by formulas over the `latest` sheet.
    g = get_column_letter(COLUMNS.index("price_brl") + 1)   # price col in latest
    c = get_column_letter(COLUMNS.index("carrier") + 1)     # carrier col in latest
    rng_p = f"latest!${g}$2:${g}$2000"
    rng_c = f"latest!${c}$2:${c}$2000"

    head_row = 8
    headers = ["carrier", "# plans", "min price", "avg price", "max price"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=j, value=h)
    _house_header(ws, head_row, len(headers))
    summ_values: dict[str, float] = {}                     # {coord: value} to bake as cached values (#21)
    for idx, carrier in enumerate(["vivo", "claro", "tim"], start=head_row + 1):
        a = f"$A${idx}"
        ws.cell(row=idx, column=1, value=carrier).font = Font(name=FONT)
        ws.cell(row=idx, column=2, value=f'=COUNTIF({rng_c},{a})').font = Font(name=FONT)
        ws.cell(row=idx, column=3, value=f'=IFERROR(_xlfn.MINIFS({rng_p},{rng_c},{a}),"-")').font = Font(name=FONT)
        ws.cell(row=idx, column=4, value=f'=IFERROR(AVERAGEIFS({rng_p},{rng_c},{a}),"-")').font = Font(name=FONT)
        ws.cell(row=idx, column=5, value=f'=IFERROR(_xlfn.MAXIFS({rng_p},{rng_c},{a}),"-")').font = Font(name=FONT)
        for col in (3, 4, 5):
            ws.cell(row=idx, column=col).number_format = CURRENCY_FMT
        # bake cached KPI values (same aggregation the formulas do, over the latest snapshot) (#21)
        prices = pd.to_numeric(latest[latest["carrier"] == carrier]["price_brl"], errors="coerce").dropna()
        summ_values[f"B{idx}"] = int((latest["carrier"] == carrier).sum())     # COUNTIF
        if len(prices):
            summ_values[f"C{idx}"] = round(float(prices.min()), 2)             # MINIFS
            summ_values[f"D{idx}"] = round(float(prices.mean()), 2)            # AVERAGEIFS
            summ_values[f"E{idx}"] = round(float(prices.max()), 2)             # MAXIFS
    for col, w in {1: 22, 2: 12, 3: 14, 4: 14, 5: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return summ_values
